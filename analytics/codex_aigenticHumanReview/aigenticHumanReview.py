from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

SCRIPT_DIR = Path(__file__).resolve().parent
ANALYTICS_DIR = SCRIPT_DIR.parent

if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))
if str(ANALYTICS_DIR) not in sys.path:
    sys.path.insert(1, str(ANALYTICS_DIR))

from review_export import export_review_items
from review_workflow import review_progress

from review_paths import decisions_json_path, input_json_path, review_dir

XLSX_PATH = ANALYTICS_DIR / "merged_master.xlsx"
WORKDIR = review_dir()
INPUT_JSON = input_json_path()
DECISIONS_JSON = decisions_json_path()
REVIEW_SHEET = "Codex Review"
REVIEW_COLUMNS = [
    "run_id", "item_id", "batch_label", "condition", "difficulty",
    "review_source_alignment", "review_distractor_quality", "review_stem_clarity",
    "review_difficulty_match", "review_decision",
    "reviewer_decision",
    "agrees_with_reviewer", "chunks_support_question", "correct_answer_verifiable",
    "distractors_clearly_wrong", "reviewer_source_call_accurate",
    "review_notes",
]


def bootstrap(refresh_input: bool, reset_decisions: bool) -> None:
    WORKDIR.mkdir(parents=True, exist_ok=True)

    exported = False
    if refresh_input or not INPUT_JSON.exists():
        if not XLSX_PATH.exists():
            raise SystemExit(f"Merged workbook not found: {XLSX_PATH}")
        count = export_review_items(XLSX_PATH, INPUT_JSON)
        exported = True
        print(f"Exported {count} review items to {INPUT_JSON}")

    if reset_decisions or not DECISIONS_JSON.exists():
        with open(DECISIONS_JSON, "w", encoding="utf-8") as f:
            json.dump([], f, indent=2)
        action = "Reset" if reset_decisions else "Initialized"
        print(f"{action} local Codex decisions at {DECISIONS_JSON}")

    if not exported and not reset_decisions:
        print(f"Codex review mirror already present: {WORKDIR}")


def show_status() -> None:
    progress = review_progress(INPUT_JSON, DECISIONS_JSON)
    total = int(progress["total"])
    decided = int(progress["decided"])
    state = str(progress["state"])
    if total:
        print(f"Codex review progress: {decided} / {total} items reviewed ({100 * decided // total}%)")
    else:
        print(f"Codex review progress: {decided} decisions ({state})")


def require_complete_status() -> None:
    progress = review_progress(INPUT_JSON, DECISIONS_JSON)
    state = str(progress["state"])
    if state == "not_exported":
        raise SystemExit(
            "Codex human review input has not been exported yet.\n"
            "Run: python codex_aigenticHumanReview\\aigenticHumanReview.py --bootstrap"
        )
    if state == "incomplete":
        total = int(progress["total"])
        decided = int(progress["decided"])
        pct = int(100 * decided / total)
        raise SystemExit(
            f"Codex human review incomplete: {decided}/{total} decisions present ({pct}%).\n"
            "Continue the manual Codex review lane and rerun --require-complete."
        )
    print(f"Codex human review complete: {progress['decided']} / {progress['total']} decisions present")


def print_paths() -> None:
    print(f"workdir={WORKDIR}")
    print(f"input={INPUT_JSON}")
    print(f"decisions={DECISIONS_JSON}")
    print(f"master={XLSX_PATH}")


def write_review_sheet() -> None:
    if not DECISIONS_JSON.exists():
        raise SystemExit(f"Codex review decisions not found: {DECISIONS_JSON}")

    with open(DECISIONS_JSON, "r", encoding="utf-8") as f:
        decisions = json.load(f)

    wb = openpyxl.load_workbook(XLSX_PATH)
    if REVIEW_SHEET in wb.sheetnames:
        del wb[REVIEW_SHEET]
    ws = wb.create_sheet(REVIEW_SHEET)

    header_fill = PatternFill("solid", fgColor="385723")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=False)

    decision_fill = {
        "ACCEPT": PatternFill("solid", fgColor="C6EFCE"),
        "REVISE": PatternFill("solid", fgColor="FFEB9C"),
        "REJECT": PatternFill("solid", fgColor="FFC7CE"),
    }
    for row_idx, d in enumerate(decisions, start=2):
        for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=d.get(col_name))
            dec = d.get("review_decision", "")
            if dec in decision_fill:
                cell.fill = decision_fill[dec]

    col_widths = {
        "run_id": 20, "item_id": 8, "batch_label": 18, "condition": 14,
        "difficulty": 10, "review_source_alignment": 22, "review_distractor_quality": 22,
        "review_stem_clarity": 18, "review_difficulty_match": 22, "review_decision": 16,
        "reviewer_decision": 18, "agrees_with_reviewer": 20, "chunks_support_question": 22,
        "correct_answer_verifiable": 24, "distractors_clearly_wrong": 22,
        "reviewer_source_call_accurate": 26, "review_notes": 60,
    }
    for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_widths.get(col_name, 15)

    ws.freeze_panes = "A2"
    wb.save(XLSX_PATH)
    print(f"Written '{REVIEW_SHEET}' sheet to {XLSX_PATH} ({len(decisions)} rows)")


def main() -> None:
    parser = argparse.ArgumentParser(description="Codex local human-review helper")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--bootstrap", action="store_true", help="Create or refresh the local Codex review mirror from merged_master.xlsx")
    group.add_argument("--status", action="store_true", help="Show local Codex review progress")
    group.add_argument("--require-complete", action="store_true", help="Exit non-zero unless the local Codex review is complete")
    group.add_argument("--print-paths", action="store_true", help="Print the active local Codex review paths")
    group.add_argument("--write", action="store_true", help="Write Codex review decisions into a Codex Review sheet in merged_master.xlsx")
    parser.add_argument("--refresh-input", action="store_true", help="Re-export review_input.json from merged_master.xlsx during --bootstrap")
    parser.add_argument("--reset-decisions", action="store_true", help="Reset codex_review_decisions.json to an empty list during --bootstrap")
    args = parser.parse_args()

    if args.bootstrap:
        bootstrap(refresh_input=args.refresh_input, reset_decisions=args.reset_decisions)
    elif args.status:
        show_status()
    elif args.require_complete:
        require_complete_status()
    elif args.print_paths:
        print_paths()
    elif args.write:
        write_review_sheet()


if __name__ == "__main__":
    main()
