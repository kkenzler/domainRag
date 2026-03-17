"""
domainRag Run Quality + Cost Dashboard
========================================
Per-batch mode:
    python viz.py [runs_dir]          # reads easy/medium/hard XLSXs from runs_dir
    python viz.py                     # default: ../_rag_testGen/runs

Merged mode:
    python viz.py --merged merged_master.xlsx
    Reads merged_master.xlsx, produces 4-condition comparison charts
    Output: analytics/merged/dashboard.png + analytics/merged/charts/

Per-batch output: analytics/<runs_dir_parent>/charts/ + dashboard.png
"""

from __future__ import annotations

import argparse
import sys
import openpyxl
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from math import pi
from pathlib import Path

# ── Theme ─────────────────────────────────────────────────────────────────────

DIFF_COLORS  = {"easy": "#4CAF50", "medium": "#FF9800", "hard": "#F44336"}
DEC_COLORS   = {"ACCEPT": "#4CAF50", "REVISE": "#FF9800", "REJECT": "#F44336", "UNKNOWN": "#546e7a"}
COND_COLORS  = {
    "local/local":  "#78909c",
    "local/haiku":  "#29b6f6",
    "haiku/local":  "#ab47bc",
    "haiku/haiku":  "#FFB300",
}
COND_ORDER   = ["local/local", "local/haiku", "haiku/local", "haiku/haiku"]

BG        = "#0d1117"
AXIS_BG   = "#161b22"
GRID_COL  = "#30363d"
TEXT_COL  = "#c9d1d9"
TITLE_COL = "#e6edf3"

plt.style.use("dark_background")


def style_ax(ax, title, ylabel=None):
    ax.set_facecolor(AXIS_BG)
    ax.tick_params(colors=TEXT_COL, labelsize=9)
    ax.xaxis.label.set_color(TEXT_COL)
    ax.yaxis.label.set_color(TEXT_COL)
    ax.set_title(title, color=TITLE_COL, fontsize=11, fontweight="bold", pad=8)
    for spine in ax.spines.values():
        spine.set_edgecolor(GRID_COL)
    ax.grid(axis="y", color=GRID_COL, linewidth=0.6, linestyle="--", alpha=0.7)
    if ylabel:
        ax.set_ylabel(ylabel, color=TEXT_COL)


# ── Cost constants (from actual llm_http.jsonl logs, hard run 50 items) ───────

_HARD_GEN_IN  = 59_475
_HARD_GEN_OUT =  9_496
_HARD_REV_IN  = 54_690
_HARD_REV_OUT =  5_367
_INGEST_IN    = 58 * (1_000 + 1_218) // 4
_INGEST_OUT   = 58 * 500 // 4

MODELS = {
    "Local\n(Qwen 7B)": (0.0,   0.0,   "#78909c"),
    "Haiku 4.5":         (0.80,  4.00,  "#29b6f6"),
    "Sonnet 4.6":        (3.00,  15.00, "#ab47bc"),
    "Opus 4.6":          (15.00, 75.00, "#ef5350"),
}


def _stage_cost(in_tok, out_tok, pi, po):
    return in_tok / 1e6 * pi + out_tok / 1e6 * po


def pipeline_costs():
    costs = {}
    for model, (pi, po, _) in MODELS.items():
        i = _stage_cost(_INGEST_IN,   _INGEST_OUT,   pi, po)
        g = _stage_cost(_HARD_GEN_IN, _HARD_GEN_OUT, pi, po)
        r = _stage_cost(_HARD_REV_IN, _HARD_REV_OUT, pi, po)
        costs[model] = {"ingest": i, "generate": g, "review": r, "total": i + g + r}
    return costs


# ─────────────────────────────────────────────────────────────────────────────
# PER-BATCH data loading
# ─────────────────────────────────────────────────────────────────────────────

def _find_runs(runs_dir: Path):
    """Auto-detect easy/medium/hard XLSXs in runs_dir."""
    found = {}
    for xlsx in sorted(runs_dir.glob("run_*.xlsx")):
        try:
            wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
            ws = wb["Items"]
            hdrs = [c.value for c in next(ws.iter_rows(max_row=1))]
            if "difficulty" not in hdrs:
                wb.close(); continue
            di = hdrs.index("difficulty")
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[di]:
                    found[str(row[di])] = xlsx
                    break
            wb.close()
        except Exception:
            pass
    order = ["easy", "medium", "hard"]
    return [(d, found[d]) for d in order if d in found]


def load_batch_run(label: str, xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    qm = {}
    for row in wb["Quality Metrics"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            qm[row[0].replace("quality.", "")] = row[1]

    hdrs = [c.value for c in next(wb["Items"].iter_rows(max_row=1))]
    idx  = {h: i for i, h in enumerate(hdrs)}
    items = []
    for row in wb["Items"].iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        items.append({
            "decision":           row[idx["decision"]],
            "source_alignment":   row[idx["source_alignment"]],
            "distractor_quality": row[idx["distractor_quality"]],
            "stem_clarity":       row[idx["stem_clarity"]],
            "difficulty_match":   row[idx["difficulty_match"]],
        })

    chunks = []
    if "Chunk Preview" in wb.sheetnames:
        for row in wb["Chunk Preview"].iter_rows(min_row=2, values_only=True):
            if row[2]:
                chunks.append(row[2])

    return {"label": label, "qm": qm, "items": items, "chunks": chunks}


# ─────────────────────────────────────────────────────────────────────────────
# PER-BATCH chart functions  (take data + labels explicitly)
# ─────────────────────────────────────────────────────────────────────────────

def chart_mean_quality(ax, data, labels):
    style_ax(ax, "Mean Quality Scores by Difficulty", "Score (1-5)")
    metrics  = ["mean_source_alignment", "mean_distractor_quality", "mean_stem_clarity"]
    m_labels = ["Source\nAlignment", "Distractor\nQuality", "Stem\nClarity"]
    x = np.arange(len(metrics))
    w = 0.25
    for i, d in enumerate(data):
        vals = [d["qm"].get(m, 0) for m in metrics]
        bars = ax.bar(x + i * w, vals, w, label=d["label"].capitalize(),
                      color=DIFF_COLORS[d["label"]], alpha=0.9, zorder=3)
        for bar, v in zip(bars, vals):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.06,
                    f"{v:.2f}", ha="center", va="bottom", color=TEXT_COL, fontsize=7.5)
    ax.set_xticks(x + w)
    ax.set_xticklabels(m_labels, fontsize=9, color=TEXT_COL)
    ax.set_ylim(0, 6)
    ax.legend(fontsize=8, framealpha=0.2, labelcolor=TEXT_COL,
              facecolor=AXIS_BG, edgecolor=GRID_COL)


def chart_decisions(ax, data, labels):
    style_ax(ax, "Reviewer Decisions by Difficulty", "Item Count")
    bottoms = np.zeros(len(data))
    for dk in ["ACCEPT", "REVISE", "REJECT"]:
        counts = [sum(1 for it in d["items"] if it["decision"] == dk) for d in data]
        bars = ax.bar(labels, counts, bottom=bottoms,
                      color=DEC_COLORS[dk], alpha=0.9, label=dk, zorder=3)
        for bar, cnt, bot in zip(bars, counts, bottoms):
            if cnt > 0:
                ax.text(bar.get_x() + bar.get_width() / 2, bot + cnt / 2,
                        str(cnt), ha="center", va="center",
                        color="white", fontsize=9, fontweight="bold")
        bottoms += np.array(counts, dtype=float)
    ax.set_ylim(0, 60)
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels([l.capitalize() for l in labels], color=TEXT_COL)
    ax.legend(fontsize=8, framealpha=0.2, labelcolor=TEXT_COL,
              facecolor=AXIS_BG, edgecolor=GRID_COL)


def chart_thresholds(ax, data, labels):
    style_ax(ax, "Threshold Pass Rates (%)", "Pass Rate (%)")
    thresh = [
        ("pct_source_alignment_gte_4",  "Src Align\n>=4"),
        ("pct_distractor_quality_gte_3","Distractor\n>=3"),
        ("pct_stem_clarity_gte_4",      "Stem Clarity\n>=4"),
        ("pct_difficulty_match_true",   "Difficulty\nMatch"),
    ]
    x = np.arange(len(thresh))
    w = 0.25
    for i, d in enumerate(data):
        vals = [d["qm"].get(k, 0) for k, _ in thresh]
        bars = ax.bar(x + i * w, vals, w, label=d["label"].capitalize(),
                      color=DIFF_COLORS[d["label"]], alpha=0.9, zorder=3)
        for bar, v in zip(bars, vals):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 1,
                    f"{v:.0f}%", ha="center", va="bottom", color=TEXT_COL, fontsize=7)
    ax.set_xticks(x + w)
    ax.set_xticklabels([lbl for _, lbl in thresh], fontsize=8, color=TEXT_COL)
    ax.set_ylim(0, 120)
    ax.legend(fontsize=8, framealpha=0.2, labelcolor=TEXT_COL,
              facecolor=AXIS_BG, edgecolor=GRID_COL)


def _boxplot(ax, data, labels, metric_key, title):
    style_ax(ax, title, "Score (1-5)")
    bp_data = [[it[metric_key] for it in d["items"] if it[metric_key] is not None]
               for d in data]
    bp = ax.boxplot(bp_data, patch_artist=True, widths=0.45,
                    medianprops=dict(color="white", linewidth=2),
                    whiskerprops=dict(color=GRID_COL), capprops=dict(color=GRID_COL),
                    flierprops=dict(marker="o", markerfacecolor=GRID_COL,
                                   markersize=4, linestyle="none"))
    for patch, lbl in zip(bp["boxes"], labels):
        patch.set_facecolor(DIFF_COLORS[lbl])
        patch.set_alpha(0.8)
    ax.set_xticklabels([l.capitalize() for l in labels], color=TEXT_COL)
    ax.set_ylim(0, 6)


def chart_radar(ax, data, labels):
    ax.set_facecolor(AXIS_BG)
    ax.set_title("Quality Radar", color=TITLE_COL, fontsize=11,
                 fontweight="bold", pad=18)
    radar = [
        ("mean_source_alignment",      "Src Align"),
        ("mean_distractor_quality",    "Distractor\nQuality"),
        ("mean_stem_clarity",          "Stem\nClarity"),
        ("pct_difficulty_match_true",  "Difficulty\nMatch"),
        ("pct_source_alignment_gte_4", "Src >=4 %"),
    ]
    def norm(key, val):
        return val / 100 if "pct" in key else (val - 1) / 4
    N = len(radar)
    angles = [n / N * 2 * pi for n in range(N)] + [0]
    ax.set_theta_offset(pi / 2)
    ax.set_theta_direction(-1)
    ax.set_rlim(0, 1)
    ax.set_rticks([0.25, 0.5, 0.75, 1.0])
    ax.set_yticklabels(["25%", "50%", "75%", "100%"], color=TEXT_COL, fontsize=7)
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels([lbl for _, lbl in radar], color=TEXT_COL, fontsize=8)
    ax.yaxis.grid(color=GRID_COL, linestyle="--", alpha=0.5)
    ax.xaxis.grid(color=GRID_COL, linestyle="--", alpha=0.5)
    ax.spines["polar"].set_color(GRID_COL)
    for d in data:
        vals = [norm(k, d["qm"].get(k, 0)) for k, _ in radar] + \
               [norm(radar[0][0], d["qm"].get(radar[0][0], 0))]
        ax.plot(angles, vals, color=DIFF_COLORS[d["label"]], linewidth=2,
                label=d["label"].capitalize())
        ax.fill(angles, vals, color=DIFF_COLORS[d["label"]], alpha=0.15)
    ax.legend(loc="upper right", bbox_to_anchor=(1.35, 1.15),
              fontsize=8, framealpha=0.2, labelcolor=TEXT_COL,
              facecolor=AXIS_BG, edgecolor=GRID_COL)


def chart_accept_vs_match(ax, data, labels):
    style_ax(ax, "Accept Rate vs Difficulty Match (%)", "%")
    accept_pct = [
        sum(1 for it in d["items"] if it["decision"] == "ACCEPT") / len(d["items"]) * 100
        for d in data
    ]
    diff_match = [d["qm"].get("pct_difficulty_match_true", 0) for d in data]
    x = np.arange(len(labels))
    w = 0.35
    b1 = ax.bar(x - w/2, accept_pct, w, color="#29b6f6", alpha=0.9,
                label="Accept Rate %", zorder=3)
    b2 = ax.bar(x + w/2, diff_match, w, color="#ce93d8", alpha=0.9,
                label="Difficulty Match %", zorder=3)
    for bar, v in list(zip(b1, accept_pct)) + list(zip(b2, diff_match)):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 1.5,
                f"{v:.0f}%", ha="center", va="bottom", color=TEXT_COL, fontsize=8)
    ax.set_xticks(x)
    ax.set_xticklabels([l.capitalize() for l in labels], color=TEXT_COL)
    ax.set_ylim(0, 115)
    ax.legend(fontsize=8, framealpha=0.2, labelcolor=TEXT_COL,
              facecolor=AXIS_BG, edgecolor=GRID_COL)


def chart_heatmap(ax, data, labels):
    ax.set_facecolor(AXIS_BG)
    ax.set_title("Mean Scores Heatmap", color=TITLE_COL, fontsize=11,
                 fontweight="bold", pad=8)
    hm_metrics = ["mean_source_alignment", "mean_distractor_quality", "mean_stem_clarity"]
    hm_labels  = ["Source\nAlignment", "Distractor\nQuality", "Stem\nClarity"]
    matrix = np.array([[d["qm"].get(m, 0) for m in hm_metrics] for d in data])
    im = ax.imshow(matrix, aspect="auto", cmap="RdYlGn", vmin=1, vmax=5)
    for i in range(len(data)):
        for j in range(len(hm_metrics)):
            v = matrix[i, j]
            ax.text(j, i, f"{v:.2f}", ha="center", va="center",
                    color="black" if 2 < v < 4.5 else "white",
                    fontsize=12, fontweight="bold")
    ax.set_xticks(range(len(hm_metrics)))
    ax.set_xticklabels(hm_labels, color=TEXT_COL, fontsize=9)
    ax.set_yticks(range(len(labels)))
    ax.set_yticklabels([l.capitalize() for l in labels], color=TEXT_COL, fontsize=10)
    cbar = ax.get_figure().colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    cbar.ax.tick_params(colors=TEXT_COL, labelsize=8)
    cbar.set_label("Score", color=TEXT_COL, fontsize=8)
    for spine in ax.spines.values():
        spine.set_edgecolor(GRID_COL)


def chart_api_cost(ax):
    ax.set_facecolor(AXIS_BG)
    ax.set_title("Estimated API Cost — 50 Items (Hard)  |  Actual token logs",
                 color=TITLE_COL, fontsize=10, fontweight="bold", pad=8)
    for spine in ax.spines.values():
        spine.set_edgecolor(GRID_COL)
    ax.grid(axis="y", color=GRID_COL, linewidth=0.6, linestyle="--", alpha=0.7)
    costs = pipeline_costs()
    mnames = list(costs.keys())
    stage_colors = {"ingest": "#546e7a", "generate": "#1565c0", "review": "#4a148c"}
    stage_labels = {"ingest": "Ingest (58 chunks)", "generate": "Generate (50 items)",
                    "review": "Review (50 items)"}
    bottoms = np.zeros(len(mnames))
    x = np.arange(len(mnames))
    for stage in ["ingest", "generate", "review"]:
        vals = np.array([costs[m][stage] for m in mnames])
        bars = ax.bar(x, vals, bottom=bottoms, color=stage_colors[stage],
                      alpha=0.9, label=stage_labels[stage], zorder=3, width=0.55)
        for bar, v, bot in zip(bars, vals, bottoms):
            if v >= 0.005:
                ax.text(bar.get_x() + bar.get_width() / 2, bot + v / 2,
                        f"${v:.3f}", ha="center", va="center",
                        color="white", fontsize=8, fontweight="bold")
        bottoms += vals
    for i, (m, bot) in enumerate(zip(mnames, bottoms)):
        lbl = "$0.00\n(local)" if bot == 0 else f"${bot:.3f}"
        ax.text(i, bot + 0.01, lbl, ha="center", va="bottom",
                color=TEXT_COL, fontsize=8.5, fontweight="bold")
    ax.set_xticks(x)
    ax.set_xticklabels(mnames, color=TEXT_COL, fontsize=9)
    ax.set_ylabel("Estimated Cost (USD)", color=TEXT_COL)
    ax.tick_params(colors=TEXT_COL, labelsize=9)
    ax.set_ylim(0, max(bottoms) * 1.25 + 0.1)
    ax.annotate(
        "Token counts from actual llm_http.jsonl logs (hard run).\n"
        "Ingest estimated: 58 chunks x ~555 in / ~125 out tokens.\n"
        "Local = $0 API cost; runtime ~70 min on Qwen 7B.",
        xy=(0.98, 0.97), xycoords="axes fraction",
        ha="right", va="top", fontsize=7.5, color="#8b949e",
        bbox=dict(boxstyle="round,pad=0.4", facecolor=AXIS_BG,
                  edgecolor=GRID_COL, alpha=0.8),
    )
    ax.legend(fontsize=8, framealpha=0.2, labelcolor=TEXT_COL,
              facecolor=AXIS_BG, edgecolor=GRID_COL, loc="upper left")


# ─────────────────────────────────────────────────────────────────────────────
# MERGED data loading
# ─────────────────────────────────────────────────────────────────────────────

def load_merged(master_path: Path) -> list:
    """
    Returns list of group dicts: {condition, difficulty, items, qm}
    One entry per (condition × difficulty) combination.
    """
    wb = openpyxl.load_workbook(master_path, data_only=True)

    # Items
    ws   = wb["Items"]
    hdrs = [c.value for c in next(ws.iter_rows(max_row=1))]
    idx  = {h: i for i, h in enumerate(hdrs)}
    groups = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        cond = row[idx.get("condition", 0)] or "unknown"
        diff = row[idx.get("difficulty", 0)] or "unknown"
        key  = (cond, diff)
        if key not in groups:
            groups[key] = {"condition": cond, "difficulty": diff, "items": [], "qm": {}}
        groups[key]["items"].append({
            "decision":           row[idx["decision"]],
            "source_alignment":   row[idx["source_alignment"]],
            "distractor_quality": row[idx["distractor_quality"]],
            "stem_clarity":       row[idx["stem_clarity"]],
            "difficulty_match":   row[idx["difficulty_match"]],
        })

    # Quality Metrics
    ws_qm   = wb["Quality Metrics"]
    qm_hdrs = [c.value for c in next(ws_qm.iter_rows(max_row=1))]
    qi      = {h: i for i, h in enumerate(qm_hdrs)}
    for row in ws_qm.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        cond = row[qi.get("condition", 0)] or "unknown"
        diff = row[qi.get("difficulty", 0)] or "unknown"
        key  = (cond, diff)
        metric = str(row[qi["metric"]]).replace("quality.", "")
        val    = row[qi["value"]]
        if key in groups and val is not None:
            # Average if multiple runs per condition/difficulty
            prev = groups[key]["qm"].get(metric)
            groups[key]["qm"][metric] = (prev + val) / 2 if prev is not None else val

    # Sort: condition order × difficulty order
    diff_order = {"easy": 0, "medium": 1, "hard": 2}
    result = sorted(
        groups.values(),
        key=lambda g: (
            COND_ORDER.index(g["condition"]) if g["condition"] in COND_ORDER else 99,
            diff_order.get(g["difficulty"], 99),
        ),
    )
    return result


def _agg_by_condition(groups: list) -> list:
    """Aggregate groups to one entry per condition (across all difficulties)."""
    from collections import defaultdict
    cond_map = defaultdict(lambda: {"items": [], "qm_lists": {}})
    for g in groups:
        c = g["condition"]
        cond_map[c]["condition"] = c
        cond_map[c]["items"].extend(g["items"])
        for k, v in g["qm"].items():
            cond_map[c]["qm_lists"].setdefault(k, []).append(v)

    result = []
    for c in COND_ORDER:
        if c not in cond_map:
            continue
        entry = cond_map[c]
        qm = {k: float(np.mean(vs)) for k, vs in entry["qm_lists"].items()}
        result.append({"condition": c, "items": entry["items"], "qm": qm})
    return result


# ─────────────────────────────────────────────────────────────────────────────
# MERGED chart functions
# ─────────────────────────────────────────────────────────────────────────────

def merged_accept_bar(ax, agg):
    """Accept rate per condition (aggregated across difficulties)."""
    style_ax(ax, "Accept Rate by Condition (all difficulties)", "Accept %")
    conds  = [g["condition"] for g in agg]
    colors = [COND_COLORS.get(c, "#aaa") for c in conds]
    rates  = [
        sum(1 for it in g["items"] if it["decision"] == "ACCEPT") / len(g["items"]) * 100
        for g in agg
    ]
    bars = ax.bar(range(len(conds)), rates, color=colors, alpha=0.9, zorder=3, width=0.6)
    for bar, v in zip(bars, rates):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 1.5,
                f"{v:.0f}%", ha="center", va="bottom", color=TEXT_COL, fontsize=10,
                fontweight="bold")
    ax.set_xticks(range(len(conds)))
    ax.set_xticklabels(conds, color=TEXT_COL, fontsize=9)
    ax.set_ylim(0, 90)


def merged_decisions(ax, agg):
    """Stacked decision distribution per condition."""
    style_ax(ax, "Reviewer Decisions by Condition", "Item Count")
    conds   = [g["condition"] for g in agg]
    bottoms = np.zeros(len(agg))
    for dk in ["ACCEPT", "REVISE", "REJECT"]:
        counts = [sum(1 for it in g["items"] if it["decision"] == dk) for g in agg]
        bars = ax.bar(range(len(conds)), counts, bottom=bottoms,
                      color=DEC_COLORS[dk], alpha=0.9, label=dk, zorder=3, width=0.6)
        for bar, cnt, bot in zip(bars, counts, bottoms):
            if cnt > 10:
                ax.text(bar.get_x() + bar.get_width() / 2, bot + cnt / 2,
                        str(cnt), ha="center", va="center",
                        color="white", fontsize=9, fontweight="bold")
        bottoms += np.array(counts, dtype=float)
    ax.set_xticks(range(len(conds)))
    ax.set_xticklabels(conds, color=TEXT_COL, fontsize=9)
    ax.set_ylim(0, max(bottoms) * 1.15)
    ax.legend(fontsize=8, framealpha=0.2, labelcolor=TEXT_COL,
              facecolor=AXIS_BG, edgecolor=GRID_COL)


def merged_accept_heatmap(ax, groups):
    """Heatmap: condition (rows) x difficulty (cols) = accept %."""
    ax.set_facecolor(AXIS_BG)
    ax.set_title("Accept Rate %  (condition x difficulty)", color=TITLE_COL,
                 fontsize=11, fontweight="bold", pad=8)
    diffs = ["easy", "medium", "hard"]
    conds = [c for c in COND_ORDER if any(g["condition"] == c for g in groups)]
    matrix = np.full((len(conds), len(diffs)), np.nan)
    for g in groups:
        ci = conds.index(g["condition"]) if g["condition"] in conds else -1
        di = diffs.index(g["difficulty"]) if g["difficulty"] in diffs else -1
        if ci >= 0 and di >= 0 and g["items"]:
            n = len(g["items"])
            matrix[ci, di] = sum(1 for it in g["items"] if it["decision"] == "ACCEPT") / n * 100
    im = ax.imshow(matrix, aspect="auto", cmap="RdYlGn", vmin=0, vmax=80)
    for i in range(len(conds)):
        for j in range(len(diffs)):
            v = matrix[i, j]
            if not np.isnan(v):
                ax.text(j, i, f"{v:.0f}%", ha="center", va="center",
                        color="black" if 20 < v < 65 else "white",
                        fontsize=13, fontweight="bold")
    ax.set_xticks(range(len(diffs)))
    ax.set_xticklabels([d.capitalize() for d in diffs], color=TEXT_COL, fontsize=10)
    ax.set_yticks(range(len(conds)))
    ax.set_yticklabels(conds, color=TEXT_COL, fontsize=9)
    cbar = ax.get_figure().colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    cbar.ax.tick_params(colors=TEXT_COL, labelsize=8)
    cbar.set_label("Accept %", color=TEXT_COL, fontsize=8)
    for spine in ax.spines.values():
        spine.set_edgecolor(GRID_COL)


def merged_quality_bar(ax, agg):
    """Mean quality scores per condition."""
    style_ax(ax, "Mean Quality Scores by Condition", "Score (1-5)")
    metrics  = ["mean_source_alignment", "mean_distractor_quality", "mean_stem_clarity"]
    m_labels = ["Source\nAlignment", "Distractor\nQuality", "Stem\nClarity"]
    x = np.arange(len(metrics))
    w = 0.18
    for i, g in enumerate(agg):
        color = COND_COLORS.get(g["condition"], "#aaa")
        vals  = [g["qm"].get(m, 0) for m in metrics]
        bars  = ax.bar(x + i * w, vals, w, label=g["condition"],
                       color=color, alpha=0.9, zorder=3)
        for bar, v in zip(bars, vals):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.05,
                    f"{v:.2f}", ha="center", va="bottom", color=TEXT_COL, fontsize=6.5)
    ax.set_xticks(x + w * 1.5)
    ax.set_xticklabels(m_labels, fontsize=9, color=TEXT_COL)
    ax.set_ylim(0, 6)
    ax.legend(fontsize=7.5, framealpha=0.2, labelcolor=TEXT_COL,
              facecolor=AXIS_BG, edgecolor=GRID_COL)


def merged_score_heatmap(ax, agg):
    """Heatmap: condition x metric = mean score."""
    ax.set_facecolor(AXIS_BG)
    ax.set_title("Mean Score Heatmap  (condition x metric)", color=TITLE_COL,
                 fontsize=11, fontweight="bold", pad=8)
    metrics   = ["mean_source_alignment", "mean_distractor_quality", "mean_stem_clarity"]
    m_labels  = ["Source\nAlignment", "Distractor\nQuality", "Stem\nClarity"]
    conds     = [g["condition"] for g in agg]
    matrix    = np.array([[g["qm"].get(m, 0) for m in metrics] for g in agg])
    im = ax.imshow(matrix, aspect="auto", cmap="RdYlGn", vmin=1, vmax=5)
    for i in range(len(agg)):
        for j in range(len(metrics)):
            v = matrix[i, j]
            ax.text(j, i, f"{v:.2f}", ha="center", va="center",
                    color="black" if 2 < v < 4.5 else "white",
                    fontsize=12, fontweight="bold")
    ax.set_xticks(range(len(metrics)))
    ax.set_xticklabels(m_labels, color=TEXT_COL, fontsize=9)
    ax.set_yticks(range(len(conds)))
    ax.set_yticklabels(conds, color=TEXT_COL, fontsize=9)
    cbar = ax.get_figure().colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    cbar.ax.tick_params(colors=TEXT_COL, labelsize=8)
    cbar.set_label("Score", color=TEXT_COL, fontsize=8)
    for spine in ax.spines.values():
        spine.set_edgecolor(GRID_COL)


def merged_radar(ax, agg):
    """Radar: one line per condition, aggregated."""
    ax.set_facecolor(AXIS_BG)
    ax.set_title("Quality Radar by Condition", color=TITLE_COL,
                 fontsize=11, fontweight="bold", pad=18)
    radar = [
        ("mean_source_alignment",      "Src Align"),
        ("mean_distractor_quality",    "Distractor"),
        ("mean_stem_clarity",          "Stem\nClarity"),
        ("pct_difficulty_match_true",  "Diff Match"),
        ("pct_source_alignment_gte_4", "Src >=4%"),
    ]
    def norm(key, val):
        return val / 100 if "pct" in key else (val - 1) / 4

    N = len(radar)
    angles = [n / N * 2 * pi for n in range(N)] + [0]
    ax.set_theta_offset(pi / 2)
    ax.set_theta_direction(-1)
    ax.set_rlim(0, 1)
    ax.set_rticks([0.25, 0.5, 0.75, 1.0])
    ax.set_yticklabels(["25%", "50%", "75%", "100%"], color=TEXT_COL, fontsize=7)
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels([lbl for _, lbl in radar], color=TEXT_COL, fontsize=8)
    ax.yaxis.grid(color=GRID_COL, linestyle="--", alpha=0.5)
    ax.xaxis.grid(color=GRID_COL, linestyle="--", alpha=0.5)
    ax.spines["polar"].set_color(GRID_COL)
    for g in agg:
        color = COND_COLORS.get(g["condition"], "#aaa")
        vals  = [norm(k, g["qm"].get(k, 0)) for k, _ in radar] + \
                [norm(radar[0][0], g["qm"].get(radar[0][0], 0))]
        ax.plot(angles, vals, color=color, linewidth=2, label=g["condition"])
        ax.fill(angles, vals, color=color, alpha=0.1)
    ax.legend(loc="upper right", bbox_to_anchor=(1.45, 1.15),
              fontsize=7.5, framealpha=0.2, labelcolor=TEXT_COL,
              facecolor=AXIS_BG, edgecolor=GRID_COL)


def merged_trend(ax, groups):
    """Line chart: accept % by difficulty, one line per condition."""
    style_ax(ax, "Accept Rate Trend: Easy -> Medium -> Hard", "Accept %")
    diffs    = ["easy", "medium", "hard"]
    x        = np.arange(len(diffs))
    conds    = [c for c in COND_ORDER if any(g["condition"] == c for g in groups)]
    plotted  = False
    for cond in conds:
        color  = COND_COLORS.get(cond, "#aaa")
        rates  = []
        for diff in diffs:
            match = [g for g in groups if g["condition"] == cond and g["difficulty"] == diff]
            if match and match[0]["items"]:
                n = len(match[0]["items"])
                rates.append(
                    sum(1 for it in match[0]["items"] if it["decision"] == "ACCEPT") / n * 100
                )
            else:
                rates.append(None)
        valid = [(xi, r) for xi, r in zip(x, rates) if r is not None]
        if valid:
            xs, ys = zip(*valid)
            ax.plot(xs, ys, color=color, linewidth=2.5, marker="o",
                    markersize=8, label=cond, zorder=3)
            for xi, yi in valid:
                ax.text(xi, yi + 2.5, f"{yi:.0f}%", ha="center",
                        color=color, fontsize=8, fontweight="bold")
            plotted = True
    ax.set_xticks(x)
    ax.set_xticklabels(["Easy", "Medium", "Hard"], color=TEXT_COL, fontsize=10)
    ax.set_ylim(0, 90)
    if plotted:
        ax.legend(fontsize=8, framealpha=0.2, labelcolor=TEXT_COL,
                  facecolor=AXIS_BG, edgecolor=GRID_COL)


def merged_diff_match(ax, agg):
    """Difficulty match % per condition."""
    style_ax(ax, "Difficulty Match % by Condition", "%")
    conds  = [g["condition"] for g in agg]
    colors = [COND_COLORS.get(c, "#aaa") for c in conds]
    vals   = [g["qm"].get("pct_difficulty_match_true", 0) for g in agg]
    bars   = ax.bar(range(len(conds)), vals, color=colors, alpha=0.9, zorder=3, width=0.6)
    for bar, v in zip(bars, vals):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 1.5,
                f"{v:.0f}%", ha="center", va="bottom", color=TEXT_COL,
                fontsize=10, fontweight="bold")
    ax.set_xticks(range(len(conds)))
    ax.set_xticklabels(conds, color=TEXT_COL, fontsize=9)
    ax.set_ylim(0, 110)


# ─────────────────────────────────────────────────────────────────────────────
# Per-batch runner
# ─────────────────────────────────────────────────────────────────────────────

def run_batch_mode(runs_dir: Path, out_dir: Path) -> None:
    charts_dir = out_dir / "charts"
    charts_dir.mkdir(exist_ok=True)

    run_files = _find_runs(runs_dir)
    if not run_files:
        print(f"No easy/medium/hard run XLSXs found in {runs_dir}")
        return

    data   = [load_batch_run(label, path) for label, path in run_files]
    labels = [d["label"] for d in data]

    CHART_SPECS = [
        ("01_mean_quality_scores",    lambda ax: chart_mean_quality(ax, data, labels),   8,   5.5, False),
        ("02_decision_distribution",  lambda ax: chart_decisions(ax, data, labels),      7,   5.5, False),
        ("03_threshold_pass_rates",   lambda ax: chart_thresholds(ax, data, labels),     9,   5.5, False),
        ("04_source_alignment_box",   lambda ax: _boxplot(ax, data, labels, "source_alignment",   "Source Alignment Distribution"),  7, 5.5, False),
        ("05_distractor_quality_box", lambda ax: _boxplot(ax, data, labels, "distractor_quality", "Distractor Quality Distribution"), 7, 5.5, False),
        ("06_stem_clarity_box",       lambda ax: _boxplot(ax, data, labels, "stem_clarity",       "Stem Clarity Distribution"),       7, 5.5, False),
        ("07_radar",                  lambda ax: chart_radar(ax, data, labels),          7,   6.0, True),
        ("08_accept_vs_match",        lambda ax: chart_accept_vs_match(ax, data, labels),7,   5.5, False),
        ("09_score_heatmap",          lambda ax: chart_heatmap(ax, data, labels),        7,   4.5, False),
        ("10_api_cost_comparison",    chart_api_cost,                                    10,  5.5, False),
    ]

    for slug, fn, w, h, polar in CHART_SPECS:
        fig, ax = plt.subplots(figsize=(w, h), subplot_kw={"polar": True} if polar else {})
        fig.patch.set_facecolor(BG)
        fn(ax)
        path = charts_dir / f"{slug}.png"
        fig.savefig(path, dpi=150, bbox_inches="tight", facecolor=BG)
        plt.close(fig)
        print(f"  chart: {path.name}")

    # Dashboard
    fig = plt.figure(figsize=(22, 22))
    fig.patch.set_facecolor(BG)
    gs = gridspec.GridSpec(4, 3, figure=fig, hspace=0.52, wspace=0.38,
                           left=0.06, right=0.97, top=0.94, bottom=0.04,
                           height_ratios=[1, 1, 1, 1.1])
    chart_mean_quality(fig.add_subplot(gs[0, 0]), data, labels)
    chart_decisions(fig.add_subplot(gs[0, 1]), data, labels)
    chart_thresholds(fig.add_subplot(gs[0, 2]), data, labels)
    _boxplot(fig.add_subplot(gs[1, 0]), data, labels, "source_alignment",   "Source Alignment Distribution")
    _boxplot(fig.add_subplot(gs[1, 1]), data, labels, "distractor_quality", "Distractor Quality Distribution")
    _boxplot(fig.add_subplot(gs[1, 2]), data, labels, "stem_clarity",       "Stem Clarity Distribution")
    chart_radar(fig.add_subplot(gs[2, 0], polar=True), data, labels)
    chart_accept_vs_match(fig.add_subplot(gs[2, 1]), data, labels)
    chart_heatmap(fig.add_subplot(gs[2, 2]), data, labels)
    chart_api_cost(fig.add_subplot(gs[3, :]))
    fig.suptitle("domainRag  —  RAG TestGen Quality & Cost Dashboard  (easy | medium | hard)",
                 color=TITLE_COL, fontsize=14, fontweight="bold", y=0.965)
    dash = out_dir / "dashboard.png"
    fig.savefig(dash, dpi=150, bbox_inches="tight", facecolor=BG)
    plt.close(fig)
    print(f"\ndashboard: {dash}")


# ─────────────────────────────────────────────────────────────────────────────
# Merged runner
# ─────────────────────────────────────────────────────────────────────────────

def run_merged_mode(master_path: Path, out_dir: Path) -> None:
    charts_dir = out_dir / "charts"
    charts_dir.mkdir(exist_ok=True)

    print(f"Loading merged data from {master_path}")
    groups = load_merged(master_path)
    agg    = _agg_by_condition(groups)
    print(f"  {len(groups)} condition/difficulty groups, {len(agg)} conditions")

    CHART_SPECS = [
        ("01_accept_by_condition",      lambda ax: merged_accept_bar(ax, agg),          8,  5.5, False),
        ("02_decisions_by_condition",   lambda ax: merged_decisions(ax, agg),           8,  5.5, False),
        ("03_accept_heatmap",           lambda ax: merged_accept_heatmap(ax, groups),   9,  5.0, False),
        ("04_quality_by_condition",     lambda ax: merged_quality_bar(ax, agg),         10, 5.5, False),
        ("05_score_heatmap",            lambda ax: merged_score_heatmap(ax, agg),       8,  4.5, False),
        ("06_radar",                    lambda ax: merged_radar(ax, agg),               7,  6.0, True),
        ("07_accept_trend",             lambda ax: merged_trend(ax, groups),            9,  5.5, False),
        ("08_difficulty_match",         lambda ax: merged_diff_match(ax, agg),          7,  5.5, False),
        ("09_api_cost_comparison",      chart_api_cost,                                 10, 5.5, False),
    ]

    for slug, fn, w, h, polar in CHART_SPECS:
        fig, ax = plt.subplots(figsize=(w, h), subplot_kw={"polar": True} if polar else {})
        fig.patch.set_facecolor(BG)
        fn(ax)
        path = charts_dir / f"{slug}.png"
        fig.savefig(path, dpi=150, bbox_inches="tight", facecolor=BG)
        plt.close(fig)
        print(f"  chart: {path.name}")

    # Dashboard 3x3
    fig = plt.figure(figsize=(22, 20))
    fig.patch.set_facecolor(BG)
    gs = gridspec.GridSpec(3, 3, figure=fig, hspace=0.52, wspace=0.38,
                           left=0.06, right=0.97, top=0.94, bottom=0.05)
    merged_accept_bar(fig.add_subplot(gs[0, 0]), agg)
    merged_decisions(fig.add_subplot(gs[0, 1]), agg)
    merged_accept_heatmap(fig.add_subplot(gs[0, 2]), groups)
    merged_quality_bar(fig.add_subplot(gs[1, 0]), agg)
    merged_score_heatmap(fig.add_subplot(gs[1, 1]), agg)
    merged_radar(fig.add_subplot(gs[1, 2], polar=True), agg)
    merged_trend(fig.add_subplot(gs[2, 0]), groups)
    merged_diff_match(fig.add_subplot(gs[2, 1]), agg)
    chart_api_cost(fig.add_subplot(gs[2, 2]))
    fig.suptitle(
        "domainRag  —  4-Condition Comparison  (local/local | local/haiku | haiku/local | haiku/haiku)",
        color=TITLE_COL, fontsize=13, fontweight="bold", y=0.965,
    )
    dash = out_dir / "dashboard.png"
    fig.savefig(dash, dpi=150, bbox_inches="tight", facecolor=BG)
    plt.close(fig)
    print(f"\ndashboard: {dash}")


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="domainRag quality dashboard")
    parser.add_argument("runs_dir", nargs="?", default=None,
                        help="Path to runs folder (per-batch mode)")
    parser.add_argument("--merged", metavar="MASTER_XLSX",
                        help="Path to merged_master.xlsx (merged mode)")
    args = parser.parse_args()

    script_dir = Path(__file__).parent

    if args.merged:
        master = Path(args.merged)
        out_dir = script_dir / "merged"
        out_dir.mkdir(exist_ok=True)
        run_merged_mode(master, out_dir)
    else:
        if args.runs_dir:
            runs_dir = Path(args.runs_dir)
        else:
            runs_dir = script_dir / "../_rag_testGen/runs"
        runs_dir = runs_dir.resolve()
        # Output alongside the runs folder's parent (for batch snapshots)
        out_dir = runs_dir.parent if runs_dir.parent != script_dir else script_dir
        run_batch_mode(runs_dir, out_dir)


if __name__ == "__main__":
    main()
