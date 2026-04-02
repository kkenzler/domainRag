"""
merge_runs.py — Merge all batch XLSX runs into a single master file.

Scans analytics/example1_local-local/ and
analytics/example1_haikuPermutations/<batch-label>/
for run_*.xlsx across all 4 batches, combines Items + Reviewer Decisions +
Traceability + Quality Metrics, adds batch_label column, writes
analytics/merged_master.xlsx.

Also includes GPT baseline workbooks from analytics/example1_gptBaseline/
if present.
GPT items are normalized to the domainRag Items schema and appended after
RAG runs. Reviewer-score columns are set to None for GPT items (no automated
reviewer was run on them). Source excerpts from GPT files are written to the
Traceability sheet so review export can surface them as context chunks.

Also runs analyticsVizs.py --merged to generate cross-condition comparison charts.

Usage:
    python merge_runs.py
"""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import openpyxl
from openpyxl import Workbook

SCRIPT_DIR = Path(__file__).resolve().parent

# ── Model selection — must match run_batches.py ───────────────────────────────
MODEL = "haiku"   # e.g. "haiku", "sonnet", "opus"

# Ordered batch label suffixes — corpus prefix is discovered automatically
BATCH_ORDER = ["local-local", f"{MODEL}-reviewer", f"{MODEL}-generator", f"{MODEL}-both"]

# Map label suffix → human-readable condition label
BATCH_LABELS = {
    "local-local":          "local/local",
    f"{MODEL}-reviewer":    f"local/{MODEL}",
    f"{MODEL}-generator":   f"{MODEL}/local",
    f"{MODEL}-both":        f"{MODEL}/{MODEL}",
}

# Sheets to merge (source sheet name → dest sheet name)
MERGE_SHEETS = ["Items", "Reviewer Decisions", "Traceability"]

# Quality Metrics: aggregate separately (one row per run)
QM_SHEET = "Quality Metrics"

# ── GPT baseline constants ────────────────────────────────────────────────────
GPT_DIR       = SCRIPT_DIR / "example1_gptBaseline"
GPT_CONDITION = "gpt/baseline"
GPT_BATCH_LABEL = "gpt-baseline"

# GPT XLSX column → canonical domainRag Items column name
_GPT_COL_MAP = {
    "item_id":        "item_id",
    "difficulty":     "difficulty",
    "question":       "question",
    "option_a":       "a",
    "option_b":       "b",
    "option_c":       "c",
    "option_d":       "d",
    "correct_answer": "correct_key",
    "source_document":"seed_doc_path",
}

# Traceability sheet canonical columns (must match cli.py output)
_TRACE_COLS = ["run_id", "item_id", "doc_path", "chunk_index", "distance", "chunk_text"]


def _batch_root(label: str) -> Path:
    """Return the directory that contains corpus-prefixed batch folders.
    local-local lives directly in analytics/; model-* live in
    analytics/example1_[MODEL]Permutations/.
    """
    if label == "local-local":
        return SCRIPT_DIR
    return SCRIPT_DIR / f"example1_{MODEL}Permutations"


def find_run_files() -> list[tuple[str, str, Path]]:
    """Returns list of (batch_label, difficulty, xlsx_path).

    Batch folders may be corpus-prefixed (e.g. example1_haiku-reviewer),
    so we glob for *_{label} rather than an exact folder name match.
    """
    runs = []
    for label in BATCH_ORDER:
        parent = _batch_root(label)
        if not parent.is_dir():
            continue
        # Glob for corpus-prefixed folders, e.g. example1_haiku-reviewer
        for batch_dir in sorted(parent.glob(f"*_{label}")):
            if not batch_dir.is_dir():
                continue
            for xlsx in sorted(batch_dir.glob("run_*.xlsx")):
                # Determine difficulty from the Items sheet
                try:
                    wb   = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
                    ws   = wb["Items"]
                    hdrs = [c.value for c in next(ws.iter_rows(max_row=1))]
                    diff_idx = hdrs.index("difficulty") if "difficulty" in hdrs else None
                    difficulty = "unknown"
                    if diff_idx is not None:
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row[0] and row[diff_idx]:
                                difficulty = str(row[diff_idx])
                                break
                    wb.close()
                except Exception:
                    difficulty = "unknown"
                runs.append((label, difficulty, xlsx))
    return runs


def merge_sheet(dest_ws, src_ws, batch_label: str, condition_label: str,
                difficulty: str, first: bool) -> None:
    """Append rows from src_ws into dest_ws. Writes header on first call."""
    rows = list(src_ws.iter_rows(values_only=True))
    if not rows:
        return

    header = list(rows[0])
    if first:
        dest_ws.append(["batch_label", "condition", "difficulty"] + header)

    for row in rows[1:]:
        if row[0] is None:
            continue
        dest_ws.append([batch_label, condition_label, difficulty] + list(row))


def merge_quality_metrics(dest_ws, src_ws, batch_label: str,
                           condition_label: str, difficulty: str,
                           run_id: str, first: bool) -> None:
    rows = list(src_ws.iter_rows(values_only=True))
    if not rows:
        return
    if first:
        dest_ws.append(["batch_label", "condition", "difficulty", "run_id", "metric", "value"])
    for row in rows[1:]:
        if row[0]:
            dest_ws.append([batch_label, condition_label, difficulty, run_id] + list(row))


def _read_gpt_xlsx(xlsx: Path) -> tuple[str, list[dict]]:
    """Read a GPT baseline XLSX and normalize to domainRag Items schema.

    Returns (difficulty, list_of_item_dicts).  Each item dict contains:
    - Mapped columns from _GPT_COL_MAP (domainRag canonical names)
    - "_source_excerpt" — raw source_excerpt text for Traceability injection
    Reviewer-score columns are absent (caller fills them as None).
    """
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    raw_hdrs = [c.value for c in next(ws.iter_rows(max_row=1))]
    src_idx = {h: i for i, h in enumerate(raw_hdrs)}

    items: list[dict] = []
    difficulty = "unknown"
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        item: dict = {}
        for src_col, dst_col in _GPT_COL_MAP.items():
            if src_col in src_idx:
                item[dst_col] = row[src_idx[src_col]]
        item["_source_excerpt"] = row[src_idx["source_excerpt"]] if "source_excerpt" in src_idx else None
        if difficulty == "unknown" and "difficulty" in src_idx and row[src_idx["difficulty"]]:
            difficulty = str(row[src_idx["difficulty"]])
        items.append(item)
    wb.close()
    return difficulty, items


def merge_gpt_baselines(
    items_ws,
    items_has_header: bool,
    trace_ws,
    trace_has_header: bool,
) -> None:
    """Append GPT baseline items into already-open merged workbook sheets.

    Reads all *.xlsx files from GPT_DIR.  For each file:
    - Normalizes columns to the domainRag Items schema
    - Appends rows to items_ws (matching the existing merged Items header)
    - Appends source excerpts to trace_ws (for review context)
    Does NOT write Reviewer Decisions or Quality Metrics rows — GPT items
    have no automated reviewer data, and callers detect this via NULL
    reviewer-score columns rather than via QM aggregates.

    Requires items_has_header=True (RAG runs must have been merged first so
    the Items header row is already in place).  Exits with a warning if the
    header is absent.
    """
    if not GPT_DIR.is_dir():
        print("\n[GPT] analytics/example1_gptBaseline/ not found — no GPT baseline files merged.")
        return

    gpt_files = sorted(GPT_DIR.glob("*.xlsx"))
    if not gpt_files:
        print("\n[GPT] No XLSX files found in analytics/example1_gptBaseline/ — skipping GPT baseline merge.")
        return

    if not items_has_header:
        print("\n[GPT] WARNING: Items sheet has no header (no RAG runs merged). "
              "GPT baseline merge requires at least one RAG run to establish the "
              "merged Items schema.  Skipping GPT merge.")
        return

    # Read merged Items header to discover column positions
    merged_hdr = [cell.value for cell in next(items_ws.iter_rows(min_row=1, max_row=1))]
    merged_idx = {h: i for i, h in enumerate(merged_hdr)}
    n_cols = len(merged_hdr)

    for xlsx in gpt_files:
        difficulty, items = _read_gpt_xlsx(xlsx)
        run_id = f"gpt-baseline-{difficulty}"

        print(f"\n  merging [{GPT_CONDITION}] {difficulty} <- {xlsx.name}  ({len(items)} items)")

        for item in items:
            # Build a None-padded row matching the merged Items header
            row: list = [None] * n_cols
            field_vals = {
                "batch_label": GPT_BATCH_LABEL,
                "condition":   GPT_CONDITION,
                "difficulty":  difficulty,
                "run_id":      run_id,
                "item_id":     item.get("item_id"),
                "mode":        "gpt",
                "question":    item.get("question"),
                "a":           item.get("a"),
                "b":           item.get("b"),
                "c":           item.get("c"),
                "d":           item.get("d"),
                "correct_key": item.get("correct_key"),
                "seed_doc_path": item.get("seed_doc_path"),
                # reviewer-score fields deliberately left None:
                # decision, source_alignment, distractor_quality,
                # stem_clarity, difficulty_match
            }
            for col_name, val in field_vals.items():
                if col_name in merged_idx:
                    row[merged_idx[col_name]] = val
            items_ws.append(row)

            # Traceability: write source_excerpt as a pseudo-chunk so review
            # export can surface it as context for the reviewer.
            excerpt = item.get("_source_excerpt")
            if trace_ws is not None and excerpt:
                if not trace_has_header:
                    trace_ws.append(_TRACE_COLS)
                    trace_has_header = True
                trace_ws.append([
                    run_id,
                    item.get("item_id"),
                    item.get("seed_doc_path"),
                    0,
                    None,   # distance: N/A for GPT (no vector retrieval)
                    excerpt,
                ])


def main() -> None:
    print("merge_runs.py — scanning batch folders\n")

    run_files = find_run_files()
    if not run_files:
        print("No run files found. Check that analytics/<batch>/ folders exist.")
        sys.exit(1)

    print(f"Found {len(run_files)} run files:")
    for label, diff, path in run_files:
        print(f"  [{BATCH_LABELS.get(label, label)}]  {diff:8s}  {path.name}")

    out_wb = Workbook()
    out_wb.remove(out_wb.active)   # remove default sheet

    # Create destination sheets
    sheet_map = {}
    for sheet_name in MERGE_SHEETS:
        ws = out_wb.create_sheet(sheet_name)
        sheet_map[sheet_name] = {"ws": ws, "first": True}

    qm_ws = out_wb.create_sheet("Quality Metrics")
    qm_first = True

    for label, difficulty, xlsx_path in run_files:
        condition = BATCH_LABELS.get(label, label)
        print(f"\n  merging [{condition}] {difficulty} <- {xlsx_path.name}")
        try:
            src_wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        except Exception as e:
            print(f"    ERROR loading {xlsx_path}: {e}")
            continue

        # Extract run_id from filename
        run_id = xlsx_path.stem.replace("run_", "")

        for sheet_name, state in sheet_map.items():
            if sheet_name in src_wb.sheetnames:
                merge_sheet(
                    state["ws"], src_wb[sheet_name],
                    label, condition, difficulty, state["first"],
                )
                state["first"] = False

        if QM_SHEET in src_wb.sheetnames:
            merge_quality_metrics(
                qm_ws, src_wb[QM_SHEET],
                label, condition, difficulty, run_id, qm_first,
            )
            qm_first = False

        src_wb.close()

    # ── GPT baseline inclusion ────────────────────────────────────────────────
    merge_gpt_baselines(
        items_ws=sheet_map["Items"]["ws"],
        items_has_header=not sheet_map["Items"]["first"],
        trace_ws=sheet_map.get("Traceability", {}).get("ws"),
        trace_has_header=not sheet_map.get("Traceability", {}).get("first", True),
    )

    out_path = SCRIPT_DIR / "merged_master.xlsx"
    out_wb.save(str(out_path))
    print(f"\nSaved: {out_path}")
    print(f"  Sheets: {out_wb.sheetnames}")

    # Trigger merged viz
    viz_py = SCRIPT_DIR / "analyticsVizs.py"
    if viz_py.exists():
        print("\nRunning analyticsVizs.py --merged ...")
        rc = subprocess.run(
            [sys.executable, str(viz_py), "--merged", str(out_path)],
            cwd=str(SCRIPT_DIR),
        ).returncode
        if rc != 0:
            print(f"  analyticsVizs.py --merged returned {rc}")

    print("\nMerge complete.")


if __name__ == "__main__":
    main()
