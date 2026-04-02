from __future__ import annotations

import json
from pathlib import Path

import openpyxl


def export_review_items(xlsx_path: Path, input_json: Path) -> int:
    """Export review items from merged_master.xlsx to the requested JSON path."""
    input_json.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    ws_items = wb["Items"]
    item_headers = [c.value for c in next(ws_items.iter_rows(min_row=1, max_row=1))]
    items = [dict(zip(item_headers, row)) for row in ws_items.iter_rows(min_row=2, values_only=True)]

    reviewer_map: dict[tuple[str, str], dict] = {}
    if "Reviewer Decisions" in wb.sheetnames:
        ws_rev = wb["Reviewer Decisions"]
        rev_headers = [c.value for c in next(ws_rev.iter_rows(min_row=1, max_row=1))]
        for row in ws_rev.iter_rows(min_row=2, values_only=True):
            d = dict(zip(rev_headers, row))
            reviewer_map[(d["run_id"], d["item_id"])] = d

    chunk_map: dict[tuple[str, str], list[dict]] = {}
    if "Traceability" in wb.sheetnames:
        ws_trace = wb["Traceability"]
        trace_headers = [c.value for c in next(ws_trace.iter_rows(min_row=1, max_row=1))]
        for row in ws_trace.iter_rows(min_row=2, values_only=True):
            d = dict(zip(trace_headers, row))
            key = (d["run_id"], d["item_id"])
            chunk_map.setdefault(key, []).append(
                {
                    "chunk_index": d["chunk_index"],
                    "distance": d["distance"],
                    "chunk_text": d["chunk_text"],
                }
            )

    wb.close()

    for key in chunk_map:
        chunk_map[key].sort(key=lambda c: (c["distance"] is None, c["distance"] or 0))

    review_items = []
    for item in items:
        key = (item.get("run_id"), item.get("item_id"))
        rev = reviewer_map.get(key, {})
        review_items.append(
            {
                "run_id": item.get("run_id"),
                "item_id": item.get("item_id"),
                "batch_label": item.get("batch_label"),
                "condition": item.get("condition"),
                "difficulty": item.get("difficulty"),
                "question": item.get("question"),
                "a": item.get("a"),
                "b": item.get("b"),
                "c": item.get("c"),
                "d": item.get("d"),
                "correct_key": item.get("correct_key"),
                "reviewer_decision": rev.get("decision"),
                "reviewer_source_alignment": rev.get("source_alignment"),
                "reviewer_distractor_quality": rev.get("distractor_quality"),
                "reviewer_stem_clarity": rev.get("stem_clarity"),
                "reviewer_difficulty_match": rev.get("difficulty_match"),
                "reviewer_reason_codes": rev.get("reason_codes"),
                "reviewer_revision_instructions": rev.get("revision_instructions"),
                "chunks": chunk_map.get(key, []),
            }
        )

    with open(input_json, "w", encoding="utf-8") as f:
        json.dump(review_items, f, indent=2, ensure_ascii=False)

    return len(review_items)
