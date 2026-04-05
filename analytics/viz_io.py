from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

import numpy as np
import openpyxl

from viz_conditions import ordered_conditions


_SHARED_INPUT_JSON = Path(__file__).resolve().parent / "review_input.json"
_SHARED_MASTER_XLSX = Path(__file__).resolve().parent / "merged_master.xlsx"


def _shared_item_metadata_map() -> dict[tuple[str, str], dict]:
    out = {}
    if _SHARED_INPUT_JSON.exists():
        with open(_SHARED_INPUT_JSON, encoding="utf-8") as f:
            rows = json.load(f)
        for row in rows:
            key = (str(row.get("run_id") or ""), str(row.get("item_id") or ""))
            out[key] = {
                "run_id": row.get("run_id"),
                "item_id": row.get("item_id"),
                "batch_label": row.get("batch_label"),
                "condition": row.get("condition"),
                "difficulty": row.get("difficulty"),
                "question": row.get("question"),
                "correct_key": row.get("correct_key"),
                "reviewer_decision": row.get("decision"),
            }
    if _SHARED_MASTER_XLSX.exists():
        wb = openpyxl.load_workbook(_SHARED_MASTER_XLSX, read_only=True, data_only=True)
        ws = wb["Items"]
        headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        idx = {h: i for i, h in enumerate(headers)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            run_id = row[idx["run_id"]] if "run_id" in idx else None
            item_id = row[idx["item_id"]] if "item_id" in idx else None
            if run_id is None or item_id is None:
                continue
            key = (str(run_id or ""), str(item_id or ""))
            base = out.setdefault(key, {})
            base.setdefault("run_id", run_id)
            base.setdefault("item_id", item_id)
            for src, dst in [
                ("batch_label", "batch_label"),
                ("condition", "condition"),
                ("difficulty", "difficulty"),
                ("question", "question"),
                ("correct_key", "correct_key"),
                ("decision", "reviewer_decision"),
            ]:
                if src in idx and base.get(dst) in {None, ""}:
                    base[dst] = row[idx[src]]
        wb.close()
    return out


def find_runs(runs_dir: Path):
    found = {}
    for xlsx in sorted(runs_dir.glob("run_*.xlsx")):
        try:
            wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
            ws = wb["Items"]
            hdrs = [c.value for c in next(ws.iter_rows(max_row=1))]
            if "difficulty" not in hdrs:
                wb.close()
                continue
            di = hdrs.index("difficulty")
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[di]:
                    found[str(row[di])] = xlsx
                    break
            wb.close()
        except Exception:
            pass
    order = ["easy", "medium", "hard"]
    return [(d, found[d]) for d in order if d in found]


def load_batch_run(label: str, xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    qm = {}
    for row in wb["Quality Metrics"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            qm[row[0].replace("quality.", "")] = row[1]

    hdrs = [c.value for c in next(wb["Items"].iter_rows(max_row=1))]
    idx = {h: i for i, h in enumerate(hdrs)}
    items = []
    for row in wb["Items"].iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        items.append(
            {
                "question": row[idx["question"]] if "question" in idx else None,
                "a": row[idx["a"]] if "a" in idx else None,
                "b": row[idx["b"]] if "b" in idx else None,
                "c": row[idx["c"]] if "c" in idx else None,
                "d": row[idx["d"]] if "d" in idx else None,
                "correct_key": row[idx["correct_key"]] if "correct_key" in idx else None,
                "seed_doc_path": row[idx["seed_doc_path"]] if "seed_doc_path" in idx else None,
                "decision": row[idx["decision"]],
                "source_alignment": row[idx["source_alignment"]],
                "distractor_quality": row[idx["distractor_quality"]],
                "stem_clarity": row[idx["stem_clarity"]],
                "difficulty_match": row[idx["difficulty_match"]],
            }
        )

    chunks = []
    if "Chunk Preview" in wb.sheetnames:
        for row in wb["Chunk Preview"].iter_rows(min_row=2, values_only=True):
            if row[2]:
                chunks.append(row[2])

    return {"label": label, "qm": qm, "items": items, "chunks": chunks}


def load_merged(master_path: Path) -> list:
    wb = openpyxl.load_workbook(master_path, data_only=True)
    ws = wb["Items"]
    hdrs = [c.value for c in next(ws.iter_rows(max_row=1))]
    idx = {h: i for i, h in enumerate(hdrs)}
    groups = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        cond = row[idx.get("condition", 0)] or "unknown"
        diff = row[idx.get("difficulty", 0)] or "unknown"
        key = (cond, diff)
        if key not in groups:
            groups[key] = {"condition": cond, "difficulty": diff, "items": [], "qm": {}, "has_reviewer_metrics": False}
        groups[key]["items"].append(
            {
                "question": row[idx["question"]] if "question" in idx else None,
                "a": row[idx["a"]] if "a" in idx else None,
                "b": row[idx["b"]] if "b" in idx else None,
                "c": row[idx["c"]] if "c" in idx else None,
                "d": row[idx["d"]] if "d" in idx else None,
                "correct_key": row[idx["correct_key"]] if "correct_key" in idx else None,
                "seed_doc_path": row[idx["seed_doc_path"]] if "seed_doc_path" in idx else None,
                "decision": row[idx["decision"]],
                "source_alignment": row[idx["source_alignment"]],
                "distractor_quality": row[idx["distractor_quality"]],
                "stem_clarity": row[idx["stem_clarity"]],
                "difficulty_match": row[idx["difficulty_match"]],
            }
        )
        if row[idx["decision"]] is not None:
            groups[key]["has_reviewer_metrics"] = True

    ws_qm = wb["Quality Metrics"]
    qm_hdrs = [c.value for c in next(ws_qm.iter_rows(max_row=1))]
    qi = {h: i for i, h in enumerate(qm_hdrs)}
    for row in ws_qm.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        cond = row[qi.get("condition", 0)] or "unknown"
        diff = row[qi.get("difficulty", 0)] or "unknown"
        key = (cond, diff)
        metric = str(row[qi["metric"]]).replace("quality.", "")
        val = row[qi["value"]]
        if key in groups and val is not None:
            prev = groups[key]["qm"].get(metric)
            groups[key]["qm"][metric] = (prev + val) / 2 if prev is not None else val

    diff_order = {"easy": 0, "medium": 1, "hard": 2}
    return sorted(
        groups.values(),
        key=lambda g: (ordered_conditions([item["condition"] for item in groups.values()]).index(g["condition"]), diff_order.get(g["difficulty"], 99)),
    )


def aggregate_by_condition(groups: list) -> list:
    cond_map = defaultdict(lambda: {"items": [], "qm_lists": {}})
    for g in groups:
        c = g["condition"]
        cond_map[c]["condition"] = c
        cond_map[c]["items"].extend(g["items"])
        for k, v in g["qm"].items():
            cond_map[c]["qm_lists"].setdefault(k, []).append(v)

    result = []
    for c in ordered_conditions(cond_map.keys()):
        if c not in cond_map:
            continue
        entry = cond_map[c]
        qm = {k: float(np.mean(vs)) for k, vs in entry["qm_lists"].items()}
        # has_reviewer_metrics is True only when at least one item in this
        # condition has a non-None decision (i.e. was run through the
        # domainRag automated reviewer).  GPT baseline items have
        # decision=None and are explicitly flagged False so that downstream
        # charts can distinguish "no reviewer data" from "score = 0".
        has_rev = any(it.get("decision") is not None for it in entry["items"])
        result.append({
            "condition": c,
            "items": entry["items"],
            "qm": qm,
            "has_reviewer_metrics": has_rev,
        })
    return result


def load_claude_review(decisions_json: Path) -> list:
    with open(decisions_json, encoding="utf-8") as f:
        raw = json.load(f)
    meta = _shared_item_metadata_map()
    return [_normalize_claude_review_item(item, meta) for item in raw]


def load_claude_review_sheet(master_path: Path) -> list:
    wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    if "Claude Review" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Claude Review"]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    items = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True) if row[0] is not None]
    wb.close()
    meta = _shared_item_metadata_map()
    return [_normalize_claude_review_item(item, meta) for item in items]


def _normalize_claude_review_item(item: dict, meta: dict[tuple[str, str], dict] | None = None) -> dict:
    key = (str(item.get("run_id") or ""), str(item.get("item_id") or ""))
    fallback = (meta or {}).get(key, {})
    item["batch_label"] = item.get("batch_label") or fallback.get("batch_label")
    item["condition"] = item.get("condition") or fallback.get("condition") or "unknown"
    item["difficulty"] = item.get("difficulty") or fallback.get("difficulty") or "unknown"
    item["question"] = item.get("question") or fallback.get("question")
    item["correct_key"] = item.get("correct_key") or fallback.get("correct_key")
    item["reviewer_decision"] = item.get("reviewer_decision") or fallback.get("reviewer_decision")
    for bfield in (
        "agrees_with_reviewer",
        "flag_ambiguity",
        "chunks_support_question",
        "correct_answer_verifiable",
        "distractors_clearly_wrong",
        "reviewer_source_call_accurate",
    ):
        v = item.get(bfield)
        if isinstance(v, str):
            item[bfield] = True if v.lower() == "true" else (False if v.lower() == "false" else v)
    for sfield in (
        "claude_source_alignment",
        "claude_distractor_quality",
        "claude_stem_clarity",
        "claude_difficulty_match",
    ):
        v = item.get(sfield)
        if isinstance(v, str) and v.strip().isdigit():
            item[sfield] = int(v.strip())
    reviewer_decision = item.get("reviewer_decision")
    lane_decision = item.get("claude_decision")
    if reviewer_decision in {None, ""}:
        item["agrees_with_reviewer"] = None
    elif item.get("agrees_with_reviewer") in {None, ""} and lane_decision not in {None, ""}:
        item["agrees_with_reviewer"] = reviewer_decision == lane_decision
    return item


def claude_review_by_condition(items: list) -> dict:
    out = defaultdict(list)
    for it in items:
        out[it["condition"]].append(it)
    return {c: out[c] for c in ordered_conditions(out.keys())}


def load_codex_review(decisions_json: Path) -> list:
    with open(decisions_json, encoding="utf-8") as f:
        raw = json.load(f)
    meta = _shared_item_metadata_map()
    return [_normalize_codex_review_item(item, meta) for item in raw]


def _normalize_codex_review_item(item: dict, meta: dict[tuple[str, str], dict] | None = None) -> dict:
    key = (str(item.get("run_id") or ""), str(item.get("item_id") or ""))
    fallback = (meta or {}).get(key, {})
    item["batch_label"] = item.get("batch_label") or fallback.get("batch_label")
    item["condition"] = item.get("condition") or fallback.get("condition") or "unknown"
    item["difficulty"] = item.get("difficulty") or fallback.get("difficulty") or "unknown"
    item["question"] = item.get("question") or fallback.get("question")
    item["correct_key"] = item.get("correct_key") or fallback.get("correct_key")
    item["reviewer_decision"] = item.get("reviewer_decision") or fallback.get("reviewer_decision")
    for bfield in (
        "agrees_with_reviewer",
        "flag_ambiguity",
        "chunks_support_question",
        "correct_answer_verifiable",
        "distractors_clearly_wrong",
        "reviewer_source_call_accurate",
    ):
        v = item.get(bfield)
        if isinstance(v, str):
            item[bfield] = True if v.lower() == "true" else (False if v.lower() == "false" else v)
    for sfield in (
        "review_source_alignment",
        "review_distractor_quality",
        "review_stem_clarity",
        "review_difficulty_match",
    ):
        v = item.get(sfield)
        if isinstance(v, str) and v.strip().isdigit():
            item[sfield] = int(v.strip())
    reviewer_decision = item.get("reviewer_decision")
    lane_decision = item.get("review_decision")
    if reviewer_decision in {None, ""}:
        item["agrees_with_reviewer"] = None
    elif item.get("agrees_with_reviewer") in {None, ""} and lane_decision not in {None, ""}:
        item["agrees_with_reviewer"] = reviewer_decision == lane_decision
    return item


def codex_review_by_condition(items: list) -> dict:
    out = defaultdict(list)
    for it in items:
        out[it["condition"]].append(it)
    return {c: out[c] for c in ordered_conditions(out.keys())}
