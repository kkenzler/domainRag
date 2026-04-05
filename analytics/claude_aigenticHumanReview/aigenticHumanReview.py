"""
aigenticHumanReview.py — Agentic Human Review helper for merged_master.xlsx

Exports items for review, optionally scores them with the configured review
provider/model, appends scored decisions, and writes the "Claude Review" sheet back to
merged_master.xlsx.

Usage:
  python aigenticHumanReview.py --export
  python aigenticHumanReview.py --score [--limit N] [--model MODEL]
  python aigenticHumanReview.py --score-opus [--limit N] [--model MODEL]
  python aigenticHumanReview.py --score-local [--limit N] [--model MODEL]
  python aigenticHumanReview.py --write
  python aigenticHumanReview.py --status
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from collections.abc import Iterable
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).parent.parent))  # analytics root — for review_workflow

from review_export import export_review_items
from review_paths import decisions_json_path, input_json_path, review_dir
from review_workflow import require_complete, review_progress

XLSX_PATH = Path(__file__).parent.parent / "merged_master.xlsx"
_REVIEW_DIR = review_dir()
INPUT_JSON = input_json_path()
DECISIONS_JSON = decisions_json_path()

REVIEW_SHEET = "Claude Review"
REVIEW_COLUMNS = [
    "run_id", "item_id", "batch_label", "condition", "difficulty",
    "claude_source_alignment", "claude_distractor_quality", "claude_stem_clarity",
    "claude_difficulty_match", "claude_decision",
    "reviewer_decision",
    "agrees_with_reviewer", "chunks_support_question", "correct_answer_verifiable",
    "distractors_clearly_wrong", "reviewer_source_call_accurate",
    "claude_notes",
]

_ANTHROPIC_URL = "https://api.anthropic.com/v1/messages"
_ANTHROPIC_VERSION = "2023-06-01"
_DEFAULT_OPUS_MODEL = "claude-opus-4-6-20250401"
_DEFAULT_LOCAL_MODEL = "qwen2.5-7b-instruct-uncensored"
_DEFAULT_LM_URL = "http://localhost:1234"
_DEFAULT_CONFIG_ENV = Path.home() / "secrets" / "domainRag" / "config.env"

_SYSTEM_PROMPT = """You are performing rigorous final-stage review for domainRag.

Assess each multiple-choice item against the supplied retrieved chunks and the
prior automated reviewer assessment. Behave like a careful expert human
reviewer.

Return exactly one JSON object and nothing else.

Required keys:
- claude_source_alignment: integer 1-5
- claude_distractor_quality: integer 1-5
- claude_stem_clarity: integer 1-5
- claude_difficulty_match: integer 1-5
- claude_decision: ACCEPT | REVISE | REJECT
- agrees_with_reviewer: boolean
- flag_ambiguity: boolean
- chunks_support_question: boolean
- correct_answer_verifiable: boolean
- distractors_clearly_wrong: boolean
- reviewer_source_call_accurate: boolean
- claude_notes: short string
"""


def _read_secret_env(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        values[k.strip()] = v.strip()
    return values


def _config_env_path() -> Path:
    override = (os.environ.get("DOMAINRAG_CONFIG_ENV") or "").strip()
    return Path(override).expanduser().resolve() if override else _DEFAULT_CONFIG_ENV


def _config_value(cfg: dict[str, str], *keys: str, default: str = "") -> str:
    for key in keys:
        val = (os.environ.get(key) or "").strip()
        if val:
            return val
        val = (cfg.get(key) or "").strip()
        if val:
            return val
    return default


def _resolve_review_runtime(model_override: str | None, provider_override: str | None) -> dict[str, str]:
    cfg = _read_secret_env(_config_env_path())
    routing = _config_value(cfg, "REVIEW_PROVIDER", default="local").lower()
    if provider_override:
        routing = provider_override.strip().lower()
    if routing not in {"local", "api"}:
        raise SystemExit(f"Invalid REVIEW_PROVIDER={routing!r}; expected 'local' or 'api'.")

    if routing == "api":
        provider = _config_value(cfg, "API_PROVIDER", default="anthropic").lower()
        model = (model_override or _config_value(cfg, "API_MODEL", default=_DEFAULT_OPUS_MODEL)).strip()
    else:
        provider = "lmstudio"
        model = (model_override or _config_value(cfg, "REVIEW_MODEL", default=_DEFAULT_LOCAL_MODEL)).strip()

    lm_url = _config_value(cfg, "LM_URL", default=_DEFAULT_LM_URL)
    return {
        "routing": routing,
        "provider": provider,
        "model": model,
        "lm_url": lm_url,
    }


def _resolve_anthropic_key() -> str:
    for key_name in ("ANTHROPIC_API_KEY", "LLM_API_KEY"):
        val = (os.environ.get(key_name) or "").strip()
        if val:
            return val
    secrets_cfg = Path.home() / "secrets" / "domainRag" / "config.env"
    cfg = _read_secret_env(secrets_cfg)
    for key_name in ("ANTHROPIC_API_KEY", "LLM_API_KEY"):
        val = (cfg.get(key_name) or "").strip()
        if val:
            return val
    raise SystemExit(
        "Anthropic API key not found.\n"
        "Set ANTHROPIC_API_KEY or LLM_API_KEY in the environment, or place it in "
        f"{secrets_cfg}."
    )


def _resolve_api_key(provider: str) -> str:
    provider = (provider or "").strip().lower()
    cfg = _read_secret_env(_config_env_path())

    if provider == "anthropic":
        return _resolve_anthropic_key()

    env_keys = {
        "openai": ("OPENAI_API_KEY", "LLM_API_KEY"),
        "gemini": ("GEMINI_API_KEY", "LLM_API_KEY"),
    }.get(provider, ("LLM_API_KEY",))
    for key_name in env_keys:
        val = (os.environ.get(key_name) or "").strip()
        if val:
            return val
        val = (cfg.get(key_name) or "").strip()
        if val:
            return val

    raise SystemExit(
        f"API key not found for provider={provider!r}.\n"
        "Set the matching provider key in the environment, or place it in "
        f"{_config_env_path()}."
    )


def _extract_json_object(text: str) -> dict:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
        cleaned = re.sub(r"\s*```$", "", cleaned)
    try:
        data = json.loads(cleaned)
    except json.JSONDecodeError:
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start == -1 or end == -1 or end <= start:
            raise
        data = json.loads(cleaned[start:end + 1])
    if not isinstance(data, dict):
        raise ValueError("Expected a JSON object response.")
    return data


def _normalize_score(value) -> int:
    score = int(value)
    return max(1, min(5, score))


def _normalize_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        raw = value.strip().lower()
        if raw in {"true", "yes", "y", "1"}:
            return True
        if raw in {"false", "no", "n", "0"}:
            return False
    raise ValueError(f"Expected boolean, got {value!r}")


def _normalize_decision(value) -> str:
    raw = str(value or "").strip().upper()
    if raw not in {"ACCEPT", "REVISE", "REJECT"}:
        raise ValueError(f"Expected ACCEPT/REVISE/REJECT, got {value!r}")
    return raw


def _review_prompt(item: dict) -> str:
    chunks = item.get("chunks") or []
    rendered_chunks = []
    for idx, chunk in enumerate(chunks[:5], start=1):
        text = str(chunk.get("chunk_text") or "").strip()
        if len(text) > 1200:
            text = text[:1200].rstrip() + " ..."
        rendered_chunks.append(f"[Chunk {idx}] distance={chunk.get('distance')}\n{text}")
    chunk_block = "\n\n".join(rendered_chunks) if rendered_chunks else "[No traceability chunks available]"

    return (
        f"Item metadata:\n"
        f"- run_id: {item.get('run_id')}\n"
        f"- item_id: {item.get('item_id')}\n"
        f"- batch_label: {item.get('batch_label')}\n"
        f"- condition: {item.get('condition')}\n"
        f"- difficulty: {item.get('difficulty')}\n\n"
        f"Question:\n{item.get('question')}\n\n"
        f"Options:\n"
        f"A. {item.get('a')}\n"
        f"B. {item.get('b')}\n"
        f"C. {item.get('c')}\n"
        f"D. {item.get('d')}\n"
        f"Correct key: {item.get('correct_key')}\n\n"
        f"Automated reviewer:\n"
        f"- reviewer_decision: {item.get('reviewer_decision')}\n"
        f"- source_alignment: {item.get('reviewer_source_alignment')}\n"
        f"- distractor_quality: {item.get('reviewer_distractor_quality')}\n"
        f"- stem_clarity: {item.get('reviewer_stem_clarity')}\n"
        f"- difficulty_match: {item.get('reviewer_difficulty_match')}\n"
        f"- reason_codes: {item.get('reviewer_reason_codes')}\n"
        f"- revision_instructions: {item.get('reviewer_revision_instructions')}\n\n"
        f"Retrieved source chunks:\n{chunk_block}\n\n"
        "Assess the item as if you were the final human reviewer."
    )


def _call_anthropic_review(item: dict, model: str, api_key: str, timeout: int) -> dict:
    import requests

    payload = {
        "model": model,
        "max_tokens": 700,
        "temperature": 0,
        "system": _SYSTEM_PROMPT,
        "messages": [{"role": "user", "content": _review_prompt(item)}],
    }
    headers = {
        "x-api-key": api_key,
        "anthropic-version": _ANTHROPIC_VERSION,
        "Content-Type": "application/json",
    }
    response = requests.post(_ANTHROPIC_URL, json=payload, headers=headers, timeout=int(timeout))
    if not response.ok:
        raise RuntimeError(f"Anthropic API HTTP {response.status_code}\n{response.text[:800]}")
    data = response.json()
    text = "\n".join(
        block.get("text", "")
        for block in data.get("content", [])
        if block.get("type") == "text"
    )
    parsed = _extract_json_object(text)
    return {
        "claude_source_alignment": _normalize_score(parsed.get("claude_source_alignment")),
        "claude_distractor_quality": _normalize_score(parsed.get("claude_distractor_quality")),
        "claude_stem_clarity": _normalize_score(parsed.get("claude_stem_clarity")),
        "claude_difficulty_match": _normalize_score(parsed.get("claude_difficulty_match")),
        "claude_decision": _normalize_decision(parsed.get("claude_decision")),
        "agrees_with_reviewer": _normalize_bool(parsed.get("agrees_with_reviewer")),
        "flag_ambiguity": _normalize_bool(parsed.get("flag_ambiguity")),
        "chunks_support_question": _normalize_bool(parsed.get("chunks_support_question")),
        "correct_answer_verifiable": _normalize_bool(parsed.get("correct_answer_verifiable")),
        "distractors_clearly_wrong": _normalize_bool(parsed.get("distractors_clearly_wrong")),
        "reviewer_source_call_accurate": _normalize_bool(parsed.get("reviewer_source_call_accurate")),
        "claude_notes": str(parsed.get("claude_notes") or "").strip(),
    }


def _call_configured_review(item: dict, provider: str, model: str, lm_url: str, timeout: int) -> dict:
    rag_dir = Path(__file__).resolve().parent.parent / "_rag_testGen"
    if str(rag_dir) not in sys.path:
        sys.path.insert(0, str(rag_dir))
    from llm_client import call_llm

    if provider == "anthropic":
        os.environ["LLM_API_KEY"] = _resolve_api_key(provider)
        text = _call_anthropic_review(item, model, os.environ["LLM_API_KEY"], timeout)
        return text

    if provider in {"openai", "gemini"}:
        os.environ["LLM_API_KEY"] = _resolve_api_key(provider)

    text = call_llm(
        lm_url=lm_url,
        model=model,
        system_prompt=_SYSTEM_PROMPT,
        user_prompt=_review_prompt(item),
        temperature=0,
        max_tokens=700,
        request_timeout_seconds=int(timeout),
        provider=provider,
    )
    parsed = _extract_json_object(text)
    return {
        "claude_source_alignment": _normalize_score(parsed.get("claude_source_alignment")),
        "claude_distractor_quality": _normalize_score(parsed.get("claude_distractor_quality")),
        "claude_stem_clarity": _normalize_score(parsed.get("claude_stem_clarity")),
        "claude_difficulty_match": _normalize_score(parsed.get("claude_difficulty_match")),
        "claude_decision": _normalize_decision(parsed.get("claude_decision")),
        "agrees_with_reviewer": _normalize_bool(parsed.get("agrees_with_reviewer")),
        "flag_ambiguity": _normalize_bool(parsed.get("flag_ambiguity")),
        "chunks_support_question": _normalize_bool(parsed.get("chunks_support_question")),
        "correct_answer_verifiable": _normalize_bool(parsed.get("correct_answer_verifiable")),
        "distractors_clearly_wrong": _normalize_bool(parsed.get("distractors_clearly_wrong")),
        "reviewer_source_call_accurate": _normalize_bool(parsed.get("reviewer_source_call_accurate")),
        "claude_notes": str(parsed.get("claude_notes") or "").strip(),
    }


def export_items() -> None:
    """Read Items + Traceability sheets, export to the canonical Claude review dir."""
    count = export_review_items(XLSX_PATH, INPUT_JSON)
    print(f"Exported {count} items to {INPUT_JSON}")


def _load_input_rows() -> list[dict]:
    if not INPUT_JSON.exists():
        return []
    with open(INPUT_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data if isinstance(data, list) else []


def _enrich_decisions_with_input(decisions: Iterable[dict]) -> list[dict]:
    source_rows = _load_input_rows()
    source_map = {(row.get("run_id"), row.get("item_id")): row for row in source_rows}
    decision_map = {(row.get("run_id"), row.get("item_id")): row for row in decisions}

    enriched: list[dict] = []
    seen: set[tuple[str | None, str | None]] = set()

    def _merge(source_row: dict | None, decision_row: dict) -> dict:
        merged = dict(source_row or {})
        for key, value in decision_row.items():
            if value is not None or key not in merged:
                merged[key] = value
        return merged

    for source_row in source_rows:
        key = (source_row.get("run_id"), source_row.get("item_id"))
        decision_row = decision_map.get(key)
        if decision_row is None:
            continue
        enriched.append(_merge(source_row, decision_row))
        seen.add(key)

    for decision_row in decisions:
        key = (decision_row.get("run_id"), decision_row.get("item_id"))
        if key in seen:
            continue
        enriched.append(_merge(source_map.get(key), decision_row))

    return enriched


def _load_decisions(enrich_with_input: bool = True) -> list[dict]:
    if not DECISIONS_JSON.exists():
        return []
    with open(DECISIONS_JSON, "r", encoding="utf-8") as f:
        decisions = json.load(f)
    if enrich_with_input:
        decisions = _enrich_decisions_with_input(decisions)
    return decisions


def repair_metadata() -> None:
    if not DECISIONS_JSON.exists():
        raise SystemExit(f"{DECISIONS_JSON} not found. Nothing to repair.")

    with open(DECISIONS_JSON, "r", encoding="utf-8") as f:
        original = json.load(f)
    repaired = _enrich_decisions_with_input(original)

    if repaired == original:
        print("Claude decisions already contain full input metadata. No repair needed.")
        return

    timestamp = time.strftime("%Y%m%d-%H%M%S")
    backup = DECISIONS_JSON.with_name(f"{DECISIONS_JSON.stem}.pre_repair_{timestamp}{DECISIONS_JSON.suffix}")
    backup.write_text(json.dumps(original, indent=2, ensure_ascii=False), encoding="utf-8")
    DECISIONS_JSON.write_text(json.dumps(repaired, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Backed up original decisions to {backup}")
    print(f"Repaired {DECISIONS_JSON} using metadata from {INPUT_JSON}")


def score_items(limit: int | None, model: str | None, timeout: int, provider_override: str | None = None) -> None:
    if not INPUT_JSON.exists():
        raise SystemExit(f"{INPUT_JSON} not found. Run --export first.")

    with open(INPUT_JSON, "r", encoding="utf-8") as f:
        items = json.load(f)

    existing = _load_decisions(enrich_with_input=False)
    existing_map = {(d["run_id"], d["item_id"]): d for d in existing}

    pending = [item for item in items if (item["run_id"], item["item_id"]) not in existing_map]
    if limit is not None:
        pending = pending[:limit]
    if not pending:
        print("No pending review items.")
        return

    runtime = _resolve_review_runtime(model, provider_override)
    DECISIONS_JSON.parent.mkdir(parents=True, exist_ok=True)

    print(
        f"Scoring {len(pending)} item(s) with "
        f"{runtime['provider']}/{runtime['model']} (routing={runtime['routing']})"
    )
    for idx, item in enumerate(pending, start=1):
        reviewed = _call_configured_review(
            item=item,
            provider=runtime["provider"],
            model=runtime["model"],
            lm_url=runtime["lm_url"],
            timeout=timeout,
        )
        decision = {
            "run_id": item["run_id"],
            "item_id": item["item_id"],
            "batch_label": item.get("batch_label"),
            "condition": item.get("condition"),
            "difficulty": item.get("difficulty"),
            "reviewer_decision": item.get("reviewer_decision"),
            **reviewed,
        }
        existing_map[(decision["run_id"], decision["item_id"])] = decision
        with open(DECISIONS_JSON, "w", encoding="utf-8") as f:
            json.dump(_enrich_decisions_with_input(existing_map.values()), f, indent=2, ensure_ascii=False)
        print(
            f"  [{idx}/{len(pending)}] "
            f"{decision['condition']} {decision['difficulty']} "
            f"item {decision['item_id']} -> {decision['claude_decision']}"
        )
        time.sleep(0.25)

    print(f"Wrote {len(existing_map)} total decision(s) to {DECISIONS_JSON}")


def score_with_opus(limit: int | None, model: str | None, timeout: int) -> None:
    score_items(limit=limit, model=model, timeout=timeout, provider_override="api")


def score_with_local(limit: int | None, model: str | None, timeout: int) -> None:
    score_items(limit=limit, model=model, timeout=timeout, provider_override="local")


def write_review_sheet() -> None:
    """Read decisions JSON and write/overwrite the Claude Review sheet."""
    if not DECISIONS_JSON.exists():
        print(f"ERROR: {DECISIONS_JSON} not found. Score items first.", file=sys.stderr)
        sys.exit(1)

    decisions = _load_decisions(enrich_with_input=True)

    print(f"Loaded {len(decisions)} decisions from {DECISIONS_JSON}")

    wb = openpyxl.load_workbook(XLSX_PATH)
    if REVIEW_SHEET in wb.sheetnames:
        del wb[REVIEW_SHEET]
    ws = wb.create_sheet(REVIEW_SHEET)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=False)

    decision_fill = {
        "ACCEPT": PatternFill("solid", fgColor="C6EFCE"),
        "REVISE": PatternFill("solid", fgColor="FFEB9C"),
        "REJECT": PatternFill("solid", fgColor="FFC7CE"),
    }
    for row_idx, d in enumerate(decisions, start=2):
        for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=d.get(col_name))
            dec = d.get("claude_decision", "")
            if dec in decision_fill:
                cell.fill = decision_fill[dec]

    col_widths = {
        "run_id": 20, "item_id": 8, "batch_label": 18, "condition": 14,
        "difficulty": 10, "claude_source_alignment": 22, "claude_distractor_quality": 22,
        "claude_stem_clarity": 18, "claude_difficulty_match": 22, "claude_decision": 16,
        "reviewer_decision": 18, "agrees_with_reviewer": 20, "chunks_support_question": 22,
        "correct_answer_verifiable": 24, "distractors_clearly_wrong": 22,
        "reviewer_source_call_accurate": 26, "claude_notes": 60,
    }
    for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_widths.get(col_name, 15)

    ws.freeze_panes = "A2"
    wb.save(XLSX_PATH)
    print(f"Written 'Claude Review' sheet to {XLSX_PATH} ({len(decisions)} rows)")


def append_batch(batch_json_path: str) -> None:
    with open(batch_json_path, "r", encoding="utf-8") as f:
        new_decisions = json.load(f)

    existing = _load_decisions(enrich_with_input=False)

    existing_map = {(d["run_id"], d["item_id"]): d for d in existing}
    for decision in new_decisions:
        existing_map[(decision["run_id"], decision["item_id"])] = decision

    DECISIONS_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(DECISIONS_JSON, "w", encoding="utf-8") as f:
        json.dump(_enrich_decisions_with_input(existing_map.values()), f, indent=2, ensure_ascii=False)

    total = None
    if INPUT_JSON.exists():
        with open(INPUT_JSON, "r", encoding="utf-8") as f2:
            total = len(json.load(f2))
    pct = f" ({100*len(existing_map)//(total or 1)}%)" if total else ""
    print(f"Appended decisions. Total: {len(existing_map)}/{total or '?'}{pct}")


def show_status() -> None:
    progress = review_progress(INPUT_JSON, DECISIONS_JSON)
    total = progress["total"]
    decided = progress["decided"]
    if total:
        print(f"Progress: {decided} / {total} items scored ({100*decided//total}%)")
    else:
        print(f"Progress: {decided} decisions (run --export first to see total)")


def require_complete_status() -> None:
    progress = require_complete(INPUT_JSON, DECISIONS_JSON)
    print(f"Claude human review complete: {progress['decided']} / {progress['total']} decisions present")


def main() -> None:
    parser = argparse.ArgumentParser(description="Claude Human Review helper")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--export", action="store_true", help="Export items to claude_review_input.json")
    group.add_argument("--score", action="store_true", help="Score pending review items using the configured review provider/model")
    group.add_argument("--score-opus", action="store_true", help="Force API-lane scoring for pending review items")
    group.add_argument("--score-local", action="store_true", help="Force local LM Studio scoring for pending review items")
    group.add_argument("--write", action="store_true", help="Write decisions JSON to Claude Review sheet")
    group.add_argument("--repair-metadata", action="store_true", help="Backfill missing review metadata from claude_review_input.json")
    group.add_argument("--status", action="store_true", help="Show scoring progress")
    group.add_argument("--append-batch", metavar="BATCH_JSON", help="Append a batch decisions JSON file to main decisions")
    group.add_argument("--require-complete", action="store_true", help="Exit non-zero unless all exported items have review decisions")
    parser.add_argument("--limit", type=int, default=None, help="Optional max number of pending items to score for --score, --score-opus, or --score-local")
    parser.add_argument("--timeout", type=int, default=120, help="API timeout seconds for scoring requests")
    parser.add_argument("--model", default=None, help="Optional model override for the chosen scoring lane")
    args = parser.parse_args()

    if args.export:
        export_items()
    elif args.score:
        score_items(args.limit, args.model, args.timeout)
    elif args.score_opus:
        score_with_opus(args.limit, args.model, args.timeout)
    elif args.score_local:
        score_with_local(args.limit, args.model, args.timeout)
    elif args.write:
        write_review_sheet()
    elif args.repair_metadata:
        repair_metadata()
    elif args.status:
        show_status()
    elif args.append_batch:
        append_batch(args.append_batch)
    elif args.require_complete:
        require_complete_status()


if __name__ == "__main__":
    main()
