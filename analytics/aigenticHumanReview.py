"""
aigenticHumanReview.py — Agentic Human Review helper for merged_master.xlsx

Exports items for review, appends scored decisions, and writes
the "Claude Review" sheet back to merged_master.xlsx.

Usage:
  python aigenticHumanReview.py --export               # dump items + chunks to claude_review_input.json
  python aigenticHumanReview.py --write                # write claude_review_decisions.json → "Claude Review" sheet
  python aigenticHumanReview.py --status               # show progress (how many decisions made so far)
"""

import argparse
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

XLSX_PATH      = Path(__file__).parent / "merged_master.xlsx"
_REVIEW_DIR    = Path(__file__).parent / "claude-review" / "claude-review-batching"
INPUT_JSON     = _REVIEW_DIR / "claude_review_input.json"
DECISIONS_JSON = _REVIEW_DIR / "claude_review_decisions.json"

REVIEW_SHEET = "Claude Review"

# Column order for the Claude Review sheet
REVIEW_COLUMNS = [
    "run_id", "item_id", "batch_label", "condition", "difficulty",
    "claude_source_alignment", "claude_distractor_quality", "claude_stem_clarity",
    "claude_difficulty_match", "claude_decision",
    "reviewer_decision",
    "agrees_with_reviewer", "chunks_support_question", "correct_answer_verifiable",
    "distractors_clearly_wrong", "reviewer_source_call_accurate",
    "claude_notes",
]


def export_items():
    """Read Items + Traceability sheets, export to claude_review_input.json."""
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)

    # --- Items sheet ---
    ws_items = wb["Items"]
    item_headers = [c.value for c in next(ws_items.iter_rows(min_row=1, max_row=1))]
    items = []
    for row in ws_items.iter_rows(min_row=2, values_only=True):
        items.append(dict(zip(item_headers, row)))

    # --- Reviewer Decisions sheet (to capture reviewer_decision per item) ---
    ws_rev = wb["Reviewer Decisions"]
    rev_headers = [c.value for c in next(ws_rev.iter_rows(min_row=1, max_row=1))]
    reviewer_map = {}  # (run_id, item_id) -> reviewer decision row
    for row in ws_rev.iter_rows(min_row=2, values_only=True):
        d = dict(zip(rev_headers, row))
        key = (d["run_id"], d["item_id"])
        reviewer_map[key] = d

    # --- Traceability sheet ---
    ws_trace = wb["Traceability"]
    trace_headers = [c.value for c in next(ws_trace.iter_rows(min_row=1, max_row=1))]
    chunk_map = {}  # (run_id, item_id) -> list of chunks
    for row in ws_trace.iter_rows(min_row=2, values_only=True):
        d = dict(zip(trace_headers, row))
        key = (d["run_id"], d["item_id"])
        chunk_map.setdefault(key, []).append({
            "chunk_index": d["chunk_index"],
            "distance": d["distance"],
            "chunk_text": d["chunk_text"],
        })

    # Sort chunks by distance ascending (closest first)
    for key in chunk_map:
        chunk_map[key].sort(key=lambda c: (c["distance"] is None, c["distance"] or 0))

    # Assemble review items
    review_items = []
    for item in items:
        key = (item["run_id"], item["item_id"])
        rev = reviewer_map.get(key, {})
        chunks = chunk_map.get(key, [])
        review_items.append({
            "run_id": item["run_id"],
            "item_id": item["item_id"],
            "batch_label": item["batch_label"],
            "condition": item["condition"],
            "difficulty": item["difficulty"],
            # Question content
            "question": item["question"],
            "a": item["a"],
            "b": item["b"],
            "c": item["c"],
            "d": item["d"],
            "correct_key": item["correct_key"],
            # Reviewer data
            "reviewer_decision": rev.get("decision"),
            "reviewer_source_alignment": rev.get("source_alignment"),
            "reviewer_distractor_quality": rev.get("distractor_quality"),
            "reviewer_stem_clarity": rev.get("stem_clarity"),
            "reviewer_difficulty_match": rev.get("difficulty_match"),
            "reviewer_reason_codes": rev.get("reason_codes"),
            "reviewer_revision_instructions": rev.get("revision_instructions"),
            # Source chunks
            "chunks": chunks,
        })

    with open(INPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(review_items, f, indent=2, ensure_ascii=False)

    print(f"Exported {len(review_items)} items to {INPUT_JSON}")
    # Summary by condition/difficulty
    from collections import Counter
    counts = Counter((r["batch_label"], r["difficulty"]) for r in review_items)
    for k in sorted(counts, key=lambda x: (x[0] or "", x[1] or "")):
        print(f"  {str(k[0]):20s} {str(k[1]):8s}: {counts[k]}")


def write_review_sheet():
    """Read claude_review_decisions.json and write/overwrite 'Claude Review' sheet."""
    if not DECISIONS_JSON.exists():
        print(f"ERROR: {DECISIONS_JSON} not found. Score items first.", file=sys.stderr)
        sys.exit(1)

    with open(DECISIONS_JSON, "r", encoding="utf-8") as f:
        decisions = json.load(f)

    print(f"Loaded {len(decisions)} decisions from {DECISIONS_JSON}")

    wb = openpyxl.load_workbook(XLSX_PATH)

    # Remove existing Claude Review sheet if present
    if REVIEW_SHEET in wb.sheetnames:
        del wb[REVIEW_SHEET]

    ws = wb.create_sheet(REVIEW_SHEET)

    # Header row
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=False)

    # Data rows
    decision_fill = {
        "ACCEPT": PatternFill("solid", fgColor="C6EFCE"),
        "REVISE": PatternFill("solid", fgColor="FFEB9C"),
        "REJECT": PatternFill("solid", fgColor="FFC7CE"),
    }

    for row_idx, d in enumerate(decisions, start=2):
        row_data = [d.get(col) for col in REVIEW_COLUMNS]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            # Color rows by decision
            dec = d.get("claude_decision", "")
            if dec in decision_fill:
                cell.fill = decision_fill[dec]

    # Column widths
    col_widths = {
        "run_id": 20, "item_id": 8, "batch_label": 18, "condition": 14,
        "difficulty": 10, "claude_source_alignment": 22, "claude_distractor_quality": 22,
        "claude_stem_clarity": 18, "claude_difficulty_match": 22, "claude_decision": 16,
        "reviewer_decision": 18, "agrees_with_reviewer": 20, "chunks_support_question": 22,
        "correct_answer_verifiable": 24, "distractors_clearly_wrong": 22,
        "reviewer_source_call_accurate": 26, "claude_notes": 60,
    }
    for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_widths.get(col_name, 15)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(XLSX_PATH)
    print(f"Written 'Claude Review' sheet to {XLSX_PATH} ({len(decisions)} rows)")


def append_batch(batch_json_path: str):
    """Append a batch of decisions (JSON array) to the main decisions file."""
    with open(batch_json_path, "r", encoding="utf-8") as f:
        new_decisions = json.load(f)

    existing = []
    if DECISIONS_JSON.exists():
        with open(DECISIONS_JSON, "r", encoding="utf-8") as f:
            existing = json.load(f)

    # Deduplicate by (run_id, item_id) — new batch wins
    existing_keys = {(d["run_id"], d["item_id"]) for d in existing}
    added = [d for d in new_decisions if (d["run_id"], d["item_id"]) not in existing_keys]
    combined = existing + added

    with open(DECISIONS_JSON, "w", encoding="utf-8") as f:
        json.dump(combined, f, indent=2, ensure_ascii=False)

    total = None
    if INPUT_JSON.exists():
        with open(INPUT_JSON, "r", encoding="utf-8") as f2:
            total = len(json.load(f2))

    pct = f" ({100*len(combined)//(total or 1)}%)" if total else ""
    print(f"Appended {len(added)} decisions. Total: {len(combined)}/{total or '?'}{pct}")


def show_status():
    """Show how many items have been decided so far."""
    total = None
    if INPUT_JSON.exists():
        with open(INPUT_JSON, "r", encoding="utf-8") as f:
            total = len(json.load(f))

    decided = 0
    if DECISIONS_JSON.exists():
        with open(DECISIONS_JSON, "r", encoding="utf-8") as f:
            decided = len(json.load(f))

    if total:
        print(f"Progress: {decided} / {total} items scored ({100*decided//total}%)")
    else:
        print(f"Progress: {decided} decisions (run --export first to see total)")


def main():
    parser = argparse.ArgumentParser(description="Claude Human Review helper")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--export", action="store_true", help="Export items to claude_review_input.json")
    group.add_argument("--write", action="store_true", help="Write decisions JSON to Claude Review sheet")
    group.add_argument("--status", action="store_true", help="Show scoring progress")
    group.add_argument("--append-batch", metavar="BATCH_JSON", help="Append a batch decisions JSON file to main decisions")
    args = parser.parse_args()

    if args.export:
        export_items()
    elif args.write:
        write_review_sheet()
    elif args.status:
        show_status()
    elif args.append_batch:
        append_batch(args.append_batch)


if __name__ == "__main__":
    main()
