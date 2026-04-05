from __future__ import annotations

import argparse
import csv
import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Iterable

import openpyxl

from review_lanes import REVIEW_LANES, shared_input_json


ANALYTICS_DIR = Path(__file__).resolve().parent
DEFAULT_MASTER_XLSX = ANALYTICS_DIR / "merged_master.xlsx"
DEFAULT_OUTPUT_DIR = ANALYTICS_DIR / "merged" / "review_analysis"

ITEM_BASE_FIELDS = [
    "run_id",
    "item_id",
    "batch_label",
    "condition",
    "difficulty",
    "mode",
    "question",
    "a",
    "b",
    "c",
    "d",
    "correct_key",
    "decision",
    "source_alignment",
    "distractor_quality",
    "stem_clarity",
    "difficulty_match",
    "schema_ok",
    "schema_violations",
    "reviewer_schema_ok",
    "reviewer_schema_violations",
    "gen_text_clean",
    "human_edited",
    "seed_doc_path",
]

LANE_PREFIX_FIELDS = [
    "present",
    "decision",
    "source_alignment",
    "distractor_quality",
    "stem_clarity",
    "difficulty_match_raw",
    "difficulty_match_pass",
    "reviewer_decision",
    "agrees_with_reviewer",
    "flag_ambiguity",
    "chunks_support_question",
    "correct_answer_verifiable",
    "distractors_clearly_wrong",
    "reviewer_source_call_accurate",
    "notes",
]


def _read_json(path: Path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _write_json(path: Path, payload) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)


def _write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _iso_mtime(path: Path) -> str | None:
    if not path.exists():
        return None
    return datetime.fromtimestamp(path.stat().st_mtime).astimezone().isoformat()


def _elapsed_hours(start_path: Path | None, end_path: Path | None) -> float | None:
    if start_path is None or end_path is None:
        return None
    if not start_path.exists() or not end_path.exists():
        return None
    seconds = end_path.stat().st_mtime - start_path.stat().st_mtime
    if seconds < 0:
        return None
    return round(seconds / 3600.0, 3)


def _boolish(value):
    if isinstance(value, bool) or value is None:
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        raw = value.strip().lower()
        if raw in {"true", "yes", "y", "1"}:
            return True
        if raw in {"false", "no", "n", "0"}:
            return False
        if raw == "":
            return None
    return value


def _difficulty_match_pass(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value >= 4
    if isinstance(value, str):
        raw = value.strip()
        if raw.isdigit():
            return int(raw) >= 4
        return _boolish(raw)
    return None


def _safe_str(value) -> str:
    return "" if value is None else str(value)


def _sortable_tuple(values: Iterable) -> tuple:
    return tuple("" if value is None else str(value) for value in values)


def _normalize_correct_key(value) -> str:
    raw = _safe_str(value).strip().upper()
    if raw in {"A", "B", "C", "D"}:
        return raw
    if raw == "":
        return "UNKNOWN"
    letters = [part.strip() for part in raw.split("|") if part.strip()]
    if len(letters) > 1:
        return "MULTI"
    return raw


def _load_items_from_master(master_xlsx: Path) -> dict[tuple[str, str], dict]:
    wb = openpyxl.load_workbook(master_xlsx, read_only=True, data_only=True)
    ws = wb["Items"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {}
    for i, name in enumerate(headers):
        if name in ITEM_BASE_FIELDS:
            idx[name] = i

    items: dict[tuple[str, str], dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        run_id = row[idx["run_id"]] if "run_id" in idx else None
        item_id = row[idx["item_id"]] if "item_id" in idx else None
        if run_id is None or item_id is None:
            continue
        item = {}
        for field in ITEM_BASE_FIELDS:
            item[field] = row[idx[field]] if field in idx else None
        items[(_safe_str(run_id), _safe_str(item_id))] = item
    wb.close()
    return items


def _load_items_from_input(input_json: Path) -> dict[tuple[str, str], dict]:
    items = {}
    for row in _read_json(input_json):
        key = (_safe_str(row.get("run_id")), _safe_str(row.get("item_id")))
        items[key] = {
            "run_id": row.get("run_id"),
            "item_id": row.get("item_id"),
            "batch_label": row.get("batch_label"),
            "condition": row.get("condition"),
            "difficulty": row.get("difficulty"),
            "mode": None,
            "question": row.get("question"),
            "a": row.get("a"),
            "b": row.get("b"),
            "c": row.get("c"),
            "d": row.get("d"),
            "correct_key": row.get("correct_key"),
            "decision": None,
            "source_alignment": row.get("reviewer_source_alignment"),
            "distractor_quality": row.get("reviewer_distractor_quality"),
            "stem_clarity": row.get("reviewer_stem_clarity"),
            "difficulty_match": row.get("reviewer_difficulty_match"),
            "schema_ok": None,
            "schema_violations": None,
            "reviewer_schema_ok": None,
            "reviewer_schema_violations": None,
            "gen_text_clean": None,
            "human_edited": None,
            "seed_doc_path": None,
        }
    return items


def _normalize_lane_item(lane_key: str, item: dict) -> dict:
    if lane_key == "claude":
        difficulty_raw = item.get("claude_difficulty_match")
        return {
            "present": True,
            "decision": item.get("claude_decision"),
            "source_alignment": item.get("claude_source_alignment"),
            "distractor_quality": item.get("claude_distractor_quality"),
            "stem_clarity": item.get("claude_stem_clarity"),
            "difficulty_match_raw": difficulty_raw,
            "difficulty_match_pass": _difficulty_match_pass(difficulty_raw),
            "reviewer_decision": item.get("reviewer_decision"),
            "agrees_with_reviewer": _boolish(item.get("agrees_with_reviewer")),
            "flag_ambiguity": _boolish(item.get("flag_ambiguity")),
            "chunks_support_question": _boolish(item.get("chunks_support_question")),
            "correct_answer_verifiable": _boolish(item.get("correct_answer_verifiable")),
            "distractors_clearly_wrong": _boolish(item.get("distractors_clearly_wrong")),
            "reviewer_source_call_accurate": _boolish(item.get("reviewer_source_call_accurate")),
            "notes": item.get("claude_notes"),
        }

    difficulty_raw = item.get("review_difficulty_match")
    return {
        "present": True,
        "decision": item.get("review_decision"),
        "source_alignment": item.get("review_source_alignment"),
        "distractor_quality": item.get("review_distractor_quality"),
        "stem_clarity": item.get("review_stem_clarity"),
        "difficulty_match_raw": difficulty_raw,
        "difficulty_match_pass": _difficulty_match_pass(difficulty_raw),
        "reviewer_decision": item.get("reviewer_decision"),
        "agrees_with_reviewer": _boolish(item.get("agrees_with_reviewer")),
        "flag_ambiguity": _boolish(item.get("flag_ambiguity")),
        "chunks_support_question": _boolish(item.get("chunks_support_question")),
        "correct_answer_verifiable": _boolish(item.get("correct_answer_verifiable")),
        "distractors_clearly_wrong": _boolish(item.get("distractors_clearly_wrong")),
        "reviewer_source_call_accurate": _boolish(item.get("reviewer_source_call_accurate")),
        "notes": item.get("review_notes"),
    }


def _load_lane_decisions(path: Path, lane_key: str) -> dict[tuple[str, str], dict]:
    if not path.exists():
        return {}
    rows = _read_json(path)
    normalized = {}
    for row in rows:
        key = (_safe_str(row.get("run_id")), _safe_str(row.get("item_id")))
        normalized[key] = _normalize_lane_item(lane_key, row)
    return normalized


def _raw_decision_count(path: Path) -> int:
    if not path.exists():
        return 0
    return len(_read_json(path))


def _duplicate_decision_rows(path: Path, lane_key: str) -> list[dict]:
    if not path.exists():
        return []
    counts: Counter[tuple[str, str]] = Counter()
    for row in _read_json(path):
        counts[(_safe_str(row.get("run_id")), _safe_str(row.get("item_id")))] += 1
    rows = []
    for (run_id, item_id), count in sorted(counts.items(), key=lambda item: _sortable_tuple(item[0])):
        if count > 1:
            rows.append(
                {
                    "lane": lane_key,
                    "run_id": run_id,
                    "item_id": item_id,
                    "duplicate_count": count,
                }
            )
    return rows


def _build_wide_rows(
    item_map: dict[tuple[str, str], dict],
    lane_maps: dict[str, dict[tuple[str, str], dict]],
) -> list[dict]:
    rows = []
    for key in sorted(item_map.keys()):
        item = dict(item_map[key])
        row = dict(item)
        row["correct_key_normalized"] = _normalize_correct_key(item.get("correct_key"))
        lane_present = []
        decisions = {}
        for lane in REVIEW_LANES:
            normalized = lane_maps.get(lane.key, {}).get(key)
            if normalized:
                if normalized.get("reviewer_decision") in {None, ""}:
                    normalized["reviewer_decision"] = item.get("decision")
                lane_decision = normalized.get("decision")
                reviewer_decision = normalized.get("reviewer_decision")
                if reviewer_decision in {None, ""}:
                    normalized["agrees_with_reviewer"] = None
                elif normalized.get("agrees_with_reviewer") in {None, ""} and lane_decision not in {None, ""}:
                    normalized["agrees_with_reviewer"] = lane_decision == reviewer_decision
            present = normalized is not None
            row[f"{lane.key}_present"] = present
            for field in LANE_PREFIX_FIELDS:
                row[f"{lane.key}_{field}"] = normalized.get(field) if normalized else None
            if present:
                lane_present.append(lane.key)
                decisions[lane.key] = normalized.get("decision")
        row["reviewed_by_any_lane"] = bool(lane_present)
        row["reviewed_lane_count"] = len(lane_present)
        row["reviewed_lanes"] = ",".join(lane_present)
        row["dual_review_complete"] = all(bool(row.get(f"{lane.key}_present")) for lane in REVIEW_LANES)
        row["lane_decision_disagreement"] = (
            len({dec for dec in decisions.values() if dec not in {None, ""}}) > 1
        )
        row["any_ambiguity_flagged"] = any(
            row.get(f"{lane.key}_flag_ambiguity") is True for lane in REVIEW_LANES
        )
        rows.append(row)
    return rows


def _build_long_rows(wide_rows: list[dict]) -> list[dict]:
    rows = []
    for wide in wide_rows:
        for lane in REVIEW_LANES:
            if not wide.get(f"{lane.key}_present"):
                continue
            rows.append(
                {
                    "lane": lane.key,
                    "lane_label": lane.label,
                    "run_id": wide.get("run_id"),
                    "item_id": wide.get("item_id"),
                    "batch_label": wide.get("batch_label"),
                    "condition": wide.get("condition"),
                    "difficulty": wide.get("difficulty"),
                    "correct_key": wide.get("correct_key"),
                    "seed_doc_path": wide.get("seed_doc_path"),
                    "lane_decision": wide.get(f"{lane.key}_decision"),
                    "lane_source_alignment": wide.get(f"{lane.key}_source_alignment"),
                    "lane_distractor_quality": wide.get(f"{lane.key}_distractor_quality"),
                    "lane_stem_clarity": wide.get(f"{lane.key}_stem_clarity"),
                    "lane_difficulty_match_raw": wide.get(f"{lane.key}_difficulty_match_raw"),
                    "lane_difficulty_match_pass": wide.get(f"{lane.key}_difficulty_match_pass"),
                    "reviewer_decision": wide.get(f"{lane.key}_reviewer_decision"),
                    "agrees_with_reviewer": wide.get(f"{lane.key}_agrees_with_reviewer"),
                    "flag_ambiguity": wide.get(f"{lane.key}_flag_ambiguity"),
                    "chunks_support_question": wide.get(f"{lane.key}_chunks_support_question"),
                    "correct_answer_verifiable": wide.get(f"{lane.key}_correct_answer_verifiable"),
                    "distractors_clearly_wrong": wide.get(f"{lane.key}_distractors_clearly_wrong"),
                    "reviewer_source_call_accurate": wide.get(f"{lane.key}_reviewer_source_call_accurate"),
                    "lane_notes": wide.get(f"{lane.key}_notes"),
                }
            )
    return rows


def _review_scope_rows(wide_rows: list[dict], scope: str) -> list[dict]:
    if scope == "all_items":
        return wide_rows
    if scope == "dual_reviewed":
        return [row for row in wide_rows if row.get("dual_review_complete")]
    return [row for row in wide_rows if row.get(f"{scope}_present")]


def _answer_key_rows(wide_rows: list[dict]) -> list[dict]:
    rows = []
    scopes = ["all_items"] + [lane.key for lane in REVIEW_LANES] + ["dual_reviewed"]
    grouping_specs = [
        ("overall", []),
        ("by_condition", ["condition"]),
        ("by_difficulty", ["difficulty"]),
        ("by_condition_difficulty", ["condition", "difficulty"]),
        ("by_run", ["run_id"]),
        ("by_run_condition_difficulty", ["run_id", "condition", "difficulty"]),
    ]
    for scope in scopes:
        scoped = _review_scope_rows(wide_rows, scope)
        for slice_type, fields in grouping_specs:
            counts: Counter[tuple] = Counter()
            for row in scoped:
                key = tuple(row.get(field) for field in fields) + (row.get("correct_key") or "unknown",)
                key = tuple(row.get(field) for field in fields) + (row.get("correct_key_normalized") or "UNKNOWN",)
                counts[key] += 1
            for key, count in sorted(counts.items(), key=lambda item: _sortable_tuple(item[0])):
                group_values = list(key[:-1])
                rows.append(
                    {
                        "scope": scope,
                        "slice_type": slice_type,
                        "run_id": group_values[fields.index("run_id")] if "run_id" in fields else "*",
                        "condition": group_values[fields.index("condition")] if "condition" in fields else "*",
                        "difficulty": group_values[fields.index("difficulty")] if "difficulty" in fields else "*",
                        "correct_key": key[-1],
                        "item_count": count,
                    }
                )
    return rows


def _condition_difficulty_rows(wide_rows: list[dict]) -> list[dict]:
    grouped: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in wide_rows:
        grouped[(row.get("condition") or "unknown", row.get("difficulty") or "unknown")].append(row)

    rows = []
    for (condition, difficulty), items in sorted(grouped.items()):
        key_counts = Counter((row.get("correct_key") or "unknown") for row in items)
        key_counts = Counter((row.get("correct_key_normalized") or "UNKNOWN") for row in items)
        rows.append(
            {
                "condition": condition,
                "difficulty": difficulty,
                "item_count": len(items),
                "distinct_seed_doc_count": len({row.get("seed_doc_path") for row in items if row.get("seed_doc_path")}),
                "codex_reviewed_count": sum(1 for row in items if row.get("codex_present")),
                "claude_reviewed_count": sum(1 for row in items if row.get("claude_present")),
                "dual_reviewed_count": sum(1 for row in items if row.get("dual_review_complete")),
                "correct_a_count": key_counts.get("A", 0),
                "correct_b_count": key_counts.get("B", 0),
                "correct_c_count": key_counts.get("C", 0),
                "correct_d_count": key_counts.get("D", 0),
                "correct_multi_count": key_counts.get("MULTI", 0),
                "correct_unknown_count": key_counts.get("UNKNOWN", 0),
            }
        )
    return rows


def _seed_doc_rows(wide_rows: list[dict]) -> list[dict]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in wide_rows:
        grouped[row.get("seed_doc_path") or "[missing_seed_doc_path]"].append(row)

    rows = []
    for seed_doc_path, items in sorted(grouped.items()):
        rows.append(
            {
                "seed_doc_path": seed_doc_path,
                "item_count": len(items),
                "distinct_run_count": len({row.get("run_id") for row in items}),
                "distinct_condition_count": len({row.get("condition") for row in items}),
                "distinct_difficulty_count": len({row.get("difficulty") for row in items}),
                "conditions": "|".join(sorted({str(row.get("condition")) for row in items if row.get("condition")})),
                "difficulties": "|".join(sorted({str(row.get("difficulty")) for row in items if row.get("difficulty")})),
                "codex_reviewed_count": sum(1 for row in items if row.get("codex_present")),
                "claude_reviewed_count": sum(1 for row in items if row.get("claude_present")),
                "dual_reviewed_count": sum(1 for row in items if row.get("dual_review_complete")),
            }
        )
    return rows


def _correct_key_anomaly_rows(wide_rows: list[dict]) -> list[dict]:
    rows = []
    for row in wide_rows:
        normalized = row.get("correct_key_normalized")
        raw = row.get("correct_key")
        if normalized not in {"MULTI", "UNKNOWN"}:
            continue
        rows.append(
            {
                "run_id": row.get("run_id"),
                "item_id": row.get("item_id"),
                "batch_label": row.get("batch_label"),
                "condition": row.get("condition"),
                "difficulty": row.get("difficulty"),
                "correct_key_raw": raw,
                "correct_key_normalized": normalized,
                "seed_doc_path": row.get("seed_doc_path"),
            }
        )
    return rows


def _failure_mode_rows(long_rows: list[dict]) -> list[dict]:
    checks = [
        ("non_accept", lambda row: row.get("lane_decision") in {"REVISE", "REJECT"}),
        ("reject", lambda row: row.get("lane_decision") == "REJECT"),
        ("revise", lambda row: row.get("lane_decision") == "REVISE"),
        ("ambiguity_flagged", lambda row: row.get("flag_ambiguity") is True),
        ("chunks_not_supportive", lambda row: row.get("chunks_support_question") is False),
        ("answer_not_verifiable", lambda row: row.get("correct_answer_verifiable") is False),
        ("distractors_not_clear", lambda row: row.get("distractors_clearly_wrong") is False),
        ("reviewer_source_call_inaccurate", lambda row: row.get("reviewer_source_call_accurate") is False),
        ("low_source_alignment", lambda row: isinstance(row.get("lane_source_alignment"), (int, float)) and row.get("lane_source_alignment") <= 3),
        ("low_distractor_quality", lambda row: isinstance(row.get("lane_distractor_quality"), (int, float)) and row.get("lane_distractor_quality") <= 3),
        ("low_stem_clarity", lambda row: isinstance(row.get("lane_stem_clarity"), (int, float)) and row.get("lane_stem_clarity") <= 3),
        ("difficulty_mismatch", lambda row: row.get("lane_difficulty_match_pass") is False),
        ("disagrees_with_reviewer", lambda row: row.get("agrees_with_reviewer") is False),
    ]

    slice_specs = [
        ("overall", []),
        ("condition", ["condition"]),
        ("difficulty", ["difficulty"]),
        ("condition_difficulty", ["condition", "difficulty"]),
    ]

    rows = []
    lane_groups: dict[str, list[dict]] = defaultdict(list)
    for row in long_rows:
        lane_groups[row["lane"]].append(row)

    for lane, lane_rows in sorted(lane_groups.items()):
        for slice_type, fields in slice_specs:
            grouped: dict[tuple, list[dict]] = defaultdict(list)
            for row in lane_rows:
                grouped[tuple(row.get(field) for field in fields)].append(row)
            for key, items in sorted(grouped.items(), key=lambda item: _sortable_tuple(item[0])):
                denom = len(items)
                if denom == 0:
                    continue
                group_map = {field: key[i] for i, field in enumerate(fields)}
                for mode_name, predicate in checks:
                    count = sum(1 for item in items if predicate(item))
                    rows.append(
                        {
                            "lane": lane,
                            "slice_type": slice_type,
                            "condition": group_map.get("condition", "*"),
                            "difficulty": group_map.get("difficulty", "*"),
                            "failure_mode": mode_name,
                            "count": count,
                            "denominator": denom,
                            "rate_percent": round(100.0 * count / denom, 3),
                        }
                    )
    return rows


def _time_cost_rows(
    wide_rows: list[dict],
    lane_maps: dict[str, dict[tuple[str, str], dict]],
    lane_paths: dict[str, Path],
    input_json: Path,
    assumed_minutes_per_item: float,
    assumed_hourly_rate: float,
    api_cost_per_item: float,
) -> list[dict]:
    total_items = len(wide_rows)
    rows = []
    for lane in REVIEW_LANES:
        decisions_path = lane_paths[lane.key]
        raw_decisions_count = _raw_decision_count(decisions_path)
        unique_decisions_count = len(lane_maps.get(lane.key, {}))
        duplicate_key_count = max(0, raw_decisions_count - unique_decisions_count)
        completion_pct = round(100.0 * unique_decisions_count / total_items, 3) if total_items else 0.0
        observed_hours = _elapsed_hours(input_json if input_json.exists() else None, decisions_path if decisions_path.exists() else None)
        estimated_review_hours = round((unique_decisions_count * assumed_minutes_per_item) / 60.0, 3)
        estimated_human_cost = round(estimated_review_hours * assumed_hourly_rate, 2)
        estimated_api_cost = round(unique_decisions_count * api_cost_per_item, 2)
        rows.append(
            {
                "lane": lane.key,
                "lane_label": lane.label,
                "items_total": total_items,
                "raw_decisions_count": raw_decisions_count,
                "unique_item_decisions_count": unique_decisions_count,
                "duplicate_key_count": duplicate_key_count,
                "completion_percent": completion_pct,
                "input_exported_at": _iso_mtime(input_json),
                "decisions_file_modified_at": _iso_mtime(decisions_path),
                "observed_window_hours": observed_hours,
                "observed_items_per_hour": round(unique_decisions_count / observed_hours, 3) if observed_hours and observed_hours > 0 else None,
                "assumed_minutes_per_item": assumed_minutes_per_item,
                "assumed_hourly_rate_usd": assumed_hourly_rate,
                "estimated_review_hours": estimated_review_hours,
                "estimated_human_cost_usd": estimated_human_cost,
                "assumed_api_cost_per_item_usd": api_cost_per_item,
                "estimated_api_cost_usd": estimated_api_cost,
                "human_minus_api_cost_usd": round(estimated_human_cost - estimated_api_cost, 2),
                "time_assessment_note": (
                    "Observed window is derived from shared input export mtime to decisions file mtime; "
                    "decision-level timestamps are not currently recorded, so this is only a coarse bound."
                ),
            }
        )
    return rows


def _manifest(
    master_xlsx: Path,
    input_json: Path,
    output_dir: Path,
    wide_rows: list[dict],
    long_rows: list[dict],
    lane_paths: dict[str, Path],
) -> dict:
    return {
        "generated_at": datetime.now().astimezone().isoformat(),
        "master_xlsx": str(master_xlsx),
        "shared_input_json": str(input_json),
        "output_dir": str(output_dir),
        "item_count": len(wide_rows),
        "lane_review_row_count": len(long_rows),
        "lane_inputs": {
            lane.key: {
                "label": lane.label,
                "decisions_json": str(lane_paths[lane.key]),
                "exists": lane_paths[lane.key].exists(),
                "modified_at": _iso_mtime(lane_paths[lane.key]),
                "raw_decisions_count": _raw_decision_count(lane_paths[lane.key]),
                "unique_item_decisions_count": len({key for key in _load_lane_decisions(lane_paths[lane.key], lane.key)}),
            }
            for lane in REVIEW_LANES
        },
    }


def build_exports(
    master_xlsx: Path,
    input_json: Path,
    output_dir: Path,
    assumed_minutes_per_item: float,
    assumed_hourly_rate: float,
    api_cost_per_item: float,
) -> None:
    if master_xlsx.exists():
        item_map = _load_items_from_master(master_xlsx)
    elif input_json.exists():
        item_map = _load_items_from_input(input_json)
    else:
        raise SystemExit(
            f"Neither {master_xlsx} nor {input_json} is available. "
            "A merged master workbook or exported review input is required."
        )

    lane_paths = {lane.key: lane.decisions_json for lane in REVIEW_LANES}
    lane_maps = {lane.key: _load_lane_decisions(lane.decisions_json, lane.key) for lane in REVIEW_LANES}

    wide_rows = _build_wide_rows(item_map, lane_maps)
    long_rows = _build_long_rows(wide_rows)
    answer_key_rows = _answer_key_rows(wide_rows)
    cond_diff_rows = _condition_difficulty_rows(wide_rows)
    seed_doc_rows = _seed_doc_rows(wide_rows)
    failure_rows = _failure_mode_rows(long_rows)
    duplicate_rows = []
    for lane in REVIEW_LANES:
        duplicate_rows.extend(_duplicate_decision_rows(lane_paths[lane.key], lane.key))
    correct_key_anomaly_rows = _correct_key_anomaly_rows(wide_rows)
    time_cost_rows = _time_cost_rows(
        wide_rows=wide_rows,
        lane_maps=lane_maps,
        lane_paths=lane_paths,
        input_json=input_json,
        assumed_minutes_per_item=assumed_minutes_per_item,
        assumed_hourly_rate=assumed_hourly_rate,
        api_cost_per_item=api_cost_per_item,
    )
    manifest = _manifest(
        master_xlsx=master_xlsx,
        input_json=input_json,
        output_dir=output_dir,
        wide_rows=wide_rows,
        long_rows=long_rows,
        lane_paths=lane_paths,
    )

    output_dir.mkdir(parents=True, exist_ok=True)

    _write_json(output_dir / "review_analysis_manifest.json", manifest)
    _write_json(output_dir / "review_item_master_wide.json", wide_rows)
    _write_json(output_dir / "review_item_lane_long.json", long_rows)

    _write_csv(
        output_dir / "review_item_master_wide.csv",
        wide_rows,
        ITEM_BASE_FIELDS
        + ["correct_key_normalized"]
        + [f"{lane.key}_{field}" for lane in REVIEW_LANES for field in LANE_PREFIX_FIELDS]
        + ["reviewed_by_any_lane", "reviewed_lane_count", "reviewed_lanes", "dual_review_complete", "lane_decision_disagreement", "any_ambiguity_flagged"],
    )
    _write_csv(
        output_dir / "review_item_lane_long.csv",
        long_rows,
        [
            "lane",
            "lane_label",
            "run_id",
            "item_id",
            "batch_label",
            "condition",
            "difficulty",
            "correct_key",
            "seed_doc_path",
            "lane_decision",
            "lane_source_alignment",
            "lane_distractor_quality",
            "lane_stem_clarity",
            "lane_difficulty_match_raw",
            "lane_difficulty_match_pass",
            "reviewer_decision",
            "agrees_with_reviewer",
            "flag_ambiguity",
            "chunks_support_question",
            "correct_answer_verifiable",
            "distractors_clearly_wrong",
            "reviewer_source_call_accurate",
            "lane_notes",
        ],
    )
    _write_csv(
        output_dir / "review_summary_answer_keys.csv",
        answer_key_rows,
        ["scope", "slice_type", "run_id", "condition", "difficulty", "correct_key", "item_count"],
    )
    _write_csv(
        output_dir / "review_summary_condition_difficulty.csv",
        cond_diff_rows,
        [
            "condition",
            "difficulty",
            "item_count",
            "distinct_seed_doc_count",
            "codex_reviewed_count",
            "claude_reviewed_count",
            "dual_reviewed_count",
            "correct_a_count",
            "correct_b_count",
            "correct_c_count",
            "correct_d_count",
            "correct_multi_count",
            "correct_unknown_count",
        ],
    )
    _write_csv(
        output_dir / "review_summary_seed_docs.csv",
        seed_doc_rows,
        [
            "seed_doc_path",
            "item_count",
            "distinct_run_count",
            "distinct_condition_count",
            "distinct_difficulty_count",
            "conditions",
            "difficulties",
            "codex_reviewed_count",
            "claude_reviewed_count",
            "dual_reviewed_count",
        ],
    )
    _write_csv(
        output_dir / "review_summary_failure_modes.csv",
        failure_rows,
        ["lane", "slice_type", "condition", "difficulty", "failure_mode", "count", "denominator", "rate_percent"],
    )
    _write_csv(
        output_dir / "review_summary_lane_duplicates.csv",
        duplicate_rows,
        ["lane", "run_id", "item_id", "duplicate_count"],
    )
    _write_csv(
        output_dir / "review_summary_correct_key_anomalies.csv",
        correct_key_anomaly_rows,
        ["run_id", "item_id", "batch_label", "condition", "difficulty", "correct_key_raw", "correct_key_normalized", "seed_doc_path"],
    )
    _write_csv(
        output_dir / "review_summary_time_cost.csv",
        time_cost_rows,
        [
            "lane",
            "lane_label",
            "items_total",
            "raw_decisions_count",
            "unique_item_decisions_count",
            "duplicate_key_count",
            "completion_percent",
            "input_exported_at",
            "decisions_file_modified_at",
            "observed_window_hours",
            "observed_items_per_hour",
            "assumed_minutes_per_item",
            "assumed_hourly_rate_usd",
            "estimated_review_hours",
            "estimated_human_cost_usd",
            "assumed_api_cost_per_item_usd",
            "estimated_api_cost_usd",
            "human_minus_api_cost_usd",
            "time_assessment_note",
        ],
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build shared review-analysis exports for Codex and Claude lanes")
    parser.add_argument("--master-xlsx", default=str(DEFAULT_MASTER_XLSX), help="Path to merged_master.xlsx")
    parser.add_argument("--input-json", default=str(shared_input_json()), help="Path to shared review_input.json")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="Directory for analysis exports")
    parser.add_argument("--minutes-per-item", type=float, default=3.0, help="Assumed average human-review minutes per item for cost/time estimates")
    parser.add_argument("--hourly-rate", type=float, default=35.0, help="Assumed loaded hourly rate in USD for cost estimates")
    parser.add_argument("--api-cost-per-item", type=float, default=0.0, help="Optional comparative API cost per item in USD")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    build_exports(
        master_xlsx=Path(args.master_xlsx).resolve(),
        input_json=Path(args.input_json).resolve(),
        output_dir=Path(args.output_dir).resolve(),
        assumed_minutes_per_item=float(args.minutes_per_item),
        assumed_hourly_rate=float(args.hourly_rate),
        api_cost_per_item=float(args.api_cost_per_item),
    )
    print(f"Review analysis exports written to {Path(args.output_dir).resolve()}")


if __name__ == "__main__":
    main()
