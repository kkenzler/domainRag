from __future__ import annotations

"""pipeline.py — Generation pipeline with human-in-the-loop checkpoints.

Two modes:
  RAG mode:     retrieve knowledge chunks from pgvector → generate → review
  Baseline mode: load docs directly → generate → review (no pgvector)

Three checkpoints (each can be skipped via config):
  CHECKPOINT_CHUNKS   — human reviews extracted knowledge chunks after ingest
  CHECKPOINT_ITEMS    — human reviews generated MCQ items before review
  CHECKPOINT_REVIEW   — human sees reviewer flags and can correct items

Checkpoints are console-based for now. interactive_run.py owns the outer loop.
"""

import json
import os
import sys
import time
import re
import psycopg

from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from openpyxl import Workbook

from embed_lmstudio import EmbedConfig, embed_texts
from ingest import IngestConfig, ingest_domain
from llm_client import call_llm
from loaders import load_document, LoadedDoc

from db_pgvector import (
    chunks_rowcount,
    ensure_schema,
    migrate_corpus_label,
    get_db_snapshot_per_doc,
    get_db_snapshot_summary,
    get_random_chunks,
    similarity_search,
    clear_corpus,
)

from text_utils import (
    clean_generator_text,
    enforce_hygiene_on_review,
    extract_first_json_obj,
    hard_trim_after_difficulty,
    validate_generator_schema,
)


# ---------------------------------------------------------------------------
# Config dataclasses
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class GenerateConfig:
    """Parameters for RAG generate runs (reads chunks from pgvector)."""
    db_dsn: str
    lm_url: str
    embed_model: str
    generator_model: str
    review_model: str
    n_items: int
    run_id: str
    prompts_dir: Path
    out_dir: Path
    # Provider fields — "lmstudio" or api_provider string
    generate_provider: str = "lmstudio"
    review_provider: str = "lmstudio"
    corpus_label: str = ""   # filters similarity search and random chunk selection to this corpus
    top_k: int = 6
    temperature_gen: float = 0.2
    temperature_review: float = 0.0
    max_tokens_gen: int = 300
    max_tokens_review: int = 600
    request_timeout_seconds: int = 600
    sleep_seconds: float = 0.0
    checkpoint_items: bool = True
    checkpoint_review: bool = True


@dataclass(frozen=True)
class BaselineConfig:
    """Parameters for baseline (no-RAG) generate runs. Loads docs directly."""
    domain_dir: Path
    lm_url: str
    generator_model: str
    review_model: str
    n_items: int
    run_id: str
    prompts_dir: Path
    out_dir: Path
    generate_provider: str = "lmstudio"
    review_provider: str = "lmstudio"
    temperature_gen: float = 0.2
    temperature_review: float = 0.0
    max_tokens_gen: int = 300
    max_tokens_review: int = 600
    request_timeout_seconds: int = 600
    sleep_seconds: float = 0.0
    checkpoint_items: bool = True
    checkpoint_review: bool = True


@dataclass(frozen=True)
class PipelineConfig:
    """Flat config for the 'pipeline' CLI command (ingest + optional generate)."""
    db_dsn: str
    domain_dir: Path
    lm_url: str
    embed_model: str
    context_model: str
    # API lane
    api_provider: str = ""
    api_model: str = ""
    # Per-step routing
    ingest_provider: str = "local"
    generate_provider: str = "local"
    review_provider: str = "local"
    corpus_label: str = ""   # derived from domain_dir basename when blank
    embedding_dim: object = None
    batch_size: int = 32
    clear_first: bool = False
    force_ingest: bool = False
    ingest_only: bool = False
    n_items: int = 5
    generator_model: str = ""
    review_model: str = ""
    run_id: str = ""
    prompts_dir: Path = Path("_prompts")
    out_dir: Path = Path("runs")
    top_k: int = 6
    sleep_seconds: float = 0.0
    checkpoint_chunks: bool = True
    checkpoint_items: bool = True
    checkpoint_review: bool = True
    ingest_delay_seconds: float = 0.0


# ---------------------------------------------------------------------------
# Human-in-the-loop checkpoints
# ---------------------------------------------------------------------------

def _checkpoint_chunks(chunks_by_doc: dict[str, list[str]]) -> dict[str, list[str]]:
    """CHECKPOINT 1: Human reviews extracted knowledge chunks.

    Shows each chunk, allows edit or skip per chunk.
    Returns approved (and possibly edited) chunks by doc.
    """
    print("\n" + "="*60, flush=True)
    print("CHECKPOINT 1: Knowledge Chunk Review", flush=True)
    print("="*60, flush=True)

    total = sum(len(v) for v in chunks_by_doc.values())
    print(f"Extracted {total} knowledge chunks across {len(chunks_by_doc)} documents.", flush=True)
    choice = input("Review chunks? (Y/N, default N): ").strip().lower()
    if choice not in {"y", "yes"}:
        print("Skipping chunk review — all chunks approved.", flush=True)
        return chunks_by_doc

    approved: dict[str, list[str]] = {}
    for doc_name, chunks in chunks_by_doc.items():
        print(f"\n  Document: {doc_name}  ({len(chunks)} chunks)", flush=True)
        approved_chunks: list[str] = []
        for i, chunk in enumerate(chunks, 1):
            print(f"\n  --- Chunk {i}/{len(chunks)} ---", flush=True)
            print(chunk, flush=True)
            print("", flush=True)
            action = input("  [A]ccept / [E]dit / [S]kip (default A): ").strip().lower()
            if action in {"s", "skip"}:
                print("  Skipped.", flush=True)
                continue
            elif action in {"e", "edit"}:
                print("  Enter revised chunk text (blank line to finish):", flush=True)
                lines = []
                while True:
                    ln = input()
                    if ln == "":
                        if lines:
                            break
                    else:
                        lines.append(ln)
                edited = "\n".join(lines).strip()
                if edited:
                    approved_chunks.append(edited)
                    print("  Saved edited chunk.", flush=True)
                else:
                    print("  Empty — skipped.", flush=True)
            else:
                approved_chunks.append(chunk)
        approved[doc_name] = approved_chunks

    total_approved = sum(len(v) for v in approved.values())
    print(f"\nCheckpoint 1 complete: {total_approved}/{total} chunks approved.", flush=True)
    return approved


def _checkpoint_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """CHECKPOINT 2: Human reviews generated MCQ items before sending to reviewer.

    Returns approved (and possibly edited) items.
    """
    print("\n" + "="*60, flush=True)
    print("CHECKPOINT 2: Generated Item Review", flush=True)
    print("="*60, flush=True)
    print(f"Generated {len(items)} items.", flush=True)
    choice = input("Review items before sending to reviewer? (Y/N, default N): ").strip().lower()
    if choice not in {"y", "yes"}:
        print("Skipping item review — all items approved.", flush=True)
        return items

    approved: list[dict[str, Any]] = []
    for i, item in enumerate(items, 1):
        print(f"\n  --- Item {i}/{len(items)} ({item.get('item_id', '')}) ---", flush=True)
        print(item.get("gen_text_clean", ""), flush=True)
        print("", flush=True)
        action = input("  [A]ccept / [E]dit / [S]kip (default A): ").strip().lower()
        if action in {"s", "skip"}:
            print("  Skipped.", flush=True)
            continue
        elif action in {"e", "edit"}:
            print("  Enter revised item text (blank line to finish):", flush=True)
            lines = []
            while True:
                ln = input()
                if ln == "":
                    if lines:
                        break
                else:
                    lines.append(ln)
            edited = "\n".join(lines).strip()
            if edited:
                item = dict(item)
                item["gen_text_clean"] = edited
                item["human_edited"] = True
                approved.append(item)
                print("  Saved edited item.", flush=True)
            else:
                print("  Empty — skipped.", flush=True)
        else:
            approved.append(item)

    print(f"\nCheckpoint 2 complete: {len(approved)}/{len(items)} items approved.", flush=True)
    return approved


def _checkpoint_review(decisions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """CHECKPOINT 3: Human sees reviewer flags and can correct items.

    Shows only flagged (REVISE/REJECT) items. Returns all decisions,
    with human corrections applied where made.
    """
    flagged = [d for d in decisions if d.get("decision") in {"REVISE", "REJECT"}]
    if not flagged:
        print("\nNo items flagged by reviewer — all accepted.", flush=True)
        return decisions

    print("\n" + "="*60, flush=True)
    print("CHECKPOINT 3: Reviewer Flag Review", flush=True)
    print("="*60, flush=True)
    print(f"{len(flagged)} item(s) flagged by reviewer (REVISE/REJECT).", flush=True)
    choice = input("Review flagged items? (Y/N, default N): ").strip().lower()
    if choice not in {"y", "yes"}:
        print("Skipping reviewer flag review.", flush=True)
        return decisions

    decisions_by_id = {d["item_id"]: d for d in decisions}

    for d in flagged:
        item_id = d.get("item_id", "")
        print(f"\n  --- {item_id} | Decision: {d.get('decision')} ---", flush=True)
        print(f"  Failure layer: {d.get('failure_layer', '')}", flush=True)
        print(f"  Reason codes: {d.get('reason_codes', [])}", flush=True)
        print(f"  Revision instructions: {d.get('revision_instructions', '')}", flush=True)
        print("", flush=True)
        action = input("  [A]ccept reviewer decision / [O]verride to ACCEPT / [E]dit item (default A): ").strip().lower()
        if action in {"o", "override"}:
            d = dict(d)
            d["decision"] = "ACCEPT"
            d["human_override"] = True
            decisions_by_id[item_id] = d
            print("  Overridden to ACCEPT.", flush=True)
        elif action in {"e", "edit"}:
            print("  Enter corrected item text (blank line to finish):", flush=True)
            lines = []
            while True:
                ln = input()
                if ln == "":
                    if lines:
                        break
                else:
                    lines.append(ln)
            edited = "\n".join(lines).strip()
            if edited:
                d = dict(d)
                d["human_corrected_text"] = edited
                d["human_override"] = True
                decisions_by_id[item_id] = d
                print("  Saved correction.", flush=True)

    result = [decisions_by_id.get(d["item_id"], d) for d in decisions]
    print(f"\nCheckpoint 3 complete.", flush=True)
    return result


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def _parse_item_fields(gen_text: str) -> dict[str, str]:
    """Extracts question/choices/correct_key/difficulty from clean generator text."""
    t = gen_text or ""
    lines = [ln.rstrip() for ln in t.splitlines()]

    def _starts_with_any(s: str, prefixes: list[str]) -> bool:
        ss = (s or "").lstrip().lower()
        return any(ss.startswith(p) for p in prefixes)

    def _grab_label_value(label: str) -> str:
        lab = label.lower()
        for ln in lines:
            s = (ln or "").lstrip()
            if s.lower().startswith(lab) and ":" in s:
                return s.split(":", 1)[1].strip()
        return ""

    def _grab_multiline_after_label(label: str) -> str:
        lab = label.lower()
        i_label = None
        for i, ln in enumerate(lines):
            s = (ln or "").lstrip()
            if s.lower().startswith(lab):
                i_label = i
                if ":" in s:
                    inline = s.split(":", 1)[1].strip()
                    if inline:
                        return inline
                break
        if i_label is None:
            return ""
        parts: list[str] = []
        terminators = ["a)", "b)", "c)", "d)", "correct", "difficulty"]
        for j in range(i_label + 1, len(lines)):
            s = (lines[j] or "").strip()
            if not s:
                if parts:
                    break
                continue
            if _starts_with_any(s, terminators):
                break
            parts.append(s)
        return " ".join(parts).strip()

    out: dict[str, str] = {}
    out["question"] = _grab_multiline_after_label("question")
    for opt in ["a)", "b)", "c)", "d)"]:
        val = ""
        for ln in lines:
            s = (ln or "").lstrip()
            if s.lower().startswith(opt):
                val = s.split(")", 1)[1].strip() if ")" in s else ""
                break
        out[opt[0]] = val
    out["correct_key"] = _grab_label_value("correct_key") or _grab_label_value("correct key")
    out["difficulty"] = (_grab_label_value("difficulty") or "").lower()
    return out


_ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _sanitize_cell(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, str):
        return _ILLEGAL_XLSX_CHARS.sub(" ", v)
    return v


def _xlsx_write_sheet(ws, headers: list[str], rows: Iterable[list[Any]]) -> None:
    ws.append(headers)
    for r in rows:
        ws.append([_sanitize_cell(v) for v in r])


def write_run_xlsx(
    out_dir: Path,
    run_id: str,
    metadata: dict[str, Any],
    db_snapshot_summary: dict[str, Any] | None = None,
    db_snapshot_per_doc: list[dict[str, Any]] | None = None,
    chunk_preview_rows: list[dict[str, Any]] | None = None,
    items_rows: list[dict[str, Any]] | None = None,
    decisions_rows: list[dict[str, Any]] | None = None,
    trace_rows: list[dict[str, Any]] | None = None,
    include_generation: bool = True,
    include_chunk_preview: bool = True,
) -> Path:
    """Creates a single XLSX per run. Generation sheets omitted in ingest-only mode."""
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_dir / f"run_{run_id}.xlsx"

    wb = Workbook()

    # Run Metadata
    ws_meta = wb.active
    ws_meta.title = "Run Metadata"
    meta_rows: list[list[Any]] = []
    for k in sorted(metadata.keys()):
        v = metadata[k]
        if isinstance(v, (dict, list)):
            v = json.dumps(v, ensure_ascii=False)
        meta_rows.append([k, v])
    _xlsx_write_sheet(ws_meta, ["key", "value"], meta_rows)

    # DB Snapshot (RAG mode only)
    if db_snapshot_summary is not None:
        ws_snap = wb.create_sheet("DB Snapshot")
        ws_snap.append(["--- Summary ---", ""])
        for k, v in sorted((db_snapshot_summary or {}).items()):
            ws_snap.append([k, ("" if v is None else v)])
        ws_snap.append(["", ""])
        ws_snap.append(["--- Per-Document Inventory ---", ""])
        ws_snap.append(["doc_path", "chunk_count", "doc_sha256", "first_created_at", "last_updated_at"])
        for row in (db_snapshot_per_doc or []):
            ws_snap.append([
                row.get("doc_path", ""),
                row.get("chunk_count", ""),
                row.get("doc_sha256", ""),
                row.get("first_created_at", "") or "",
                row.get("last_updated_at", "") or "",
            ])

    # Chunk Preview
    if include_chunk_preview and chunk_preview_rows:
        ws_cp = wb.create_sheet("Chunk Preview")
        cp_headers = ["doc_name", "chunk_index", "chunk_chars", "chunk_text"]
        cp_data: list[list[Any]] = []
        for r in chunk_preview_rows:
            cp_data.append([
                r.get("doc_name"),
                r.get("chunk_index"),
                r.get("chunk_chars"),
                r.get("chunk_text"),
            ])
        _xlsx_write_sheet(ws_cp, cp_headers, cp_data)

    # Generation sheets
    if include_generation:
        ws_items = wb.create_sheet("Items")
        items_headers = [
            "run_id", "item_id", "mode",
            "question", "a", "b", "c", "d",
            "correct_key", "difficulty",
            "decision",
            "source_alignment", "distractor_quality", "stem_clarity", "difficulty_match",
            "schema_ok", "schema_violations",
            "reviewer_schema_ok", "reviewer_schema_violations",
            "gen_text_clean", "human_edited", "seed_doc_path",
        ]
        items_data: list[list[Any]] = []
        for r in (items_rows or []):
            items_data.append([
                r.get("run_id"), r.get("item_id"), r.get("mode", "rag"),
                r.get("question"), r.get("a"), r.get("b"), r.get("c"), r.get("d"),
                r.get("correct_key"), r.get("difficulty"),
                r.get("decision"),
                r.get("source_alignment"), r.get("distractor_quality"),
                r.get("stem_clarity"), r.get("difficulty_match"),
                r.get("schema_ok"), r.get("schema_violations"),
                r.get("reviewer_schema_ok"), r.get("reviewer_schema_violations"),
                r.get("gen_text_clean"),
                r.get("human_edited", False),
                r.get("seed_doc_path", ""),
            ])
        _xlsx_write_sheet(ws_items, items_headers, items_data)

        ws_rev = wb.create_sheet("Reviewer Decisions")
        rev_headers = [
            "run_id", "item_id", "decision",
            "source_alignment", "distractor_quality", "stem_clarity", "difficulty_match",
            "failure_layer", "reason_codes", "revision_instructions",
            "reviewer_schema_ok", "reviewer_schema_violations", "reviewer_parse_ok",
            "human_override", "human_corrected_text",
        ]
        rev_data: list[list[Any]] = []
        for r in (decisions_rows or []):
            rev_data.append([
                r.get("run_id"), r.get("item_id"), r.get("decision"),
                r.get("source_alignment"), r.get("distractor_quality"),
                r.get("stem_clarity"), r.get("difficulty_match"),
                r.get("failure_layer"),
                json.dumps(r.get("reason_codes", []), ensure_ascii=False),
                r.get("revision_instructions"),
                r.get("reviewer_schema_ok"),
                json.dumps(r.get("reviewer_schema_violations", []), ensure_ascii=False),
                r.get("reviewer_parse_ok"),
                r.get("human_override", False),
                r.get("human_corrected_text", ""),
            ])
        _xlsx_write_sheet(ws_rev, rev_headers, rev_data)

        if trace_rows:
            ws_trace = wb.create_sheet("Traceability")
            _xlsx_write_sheet(
                ws_trace,
                ["run_id", "item_id", "doc_path", "chunk_index", "distance", "chunk_text"],
                [[r.get("run_id"), r.get("item_id"), r.get("doc_path"),
                  r.get("chunk_index"), r.get("distance"), r.get("chunk_text")]
                 for r in trace_rows],
            )

        ws_q = wb.create_sheet("Quality Metrics")
        ws_q.append(["metric", "value"])
        for k in sorted(metadata.keys()):
            if str(k).startswith("quality."):
                v = metadata[k]
                if isinstance(v, (dict, list)):
                    v = json.dumps(v, ensure_ascii=False)
                ws_q.append([k, ("" if v is None else v)])

    wb.save(xlsx_path)
    return xlsx_path


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _redact_dsn(dsn: str) -> str:
    if not dsn:
        return ""
    try:
        if "://" in dsn and "@" in dsn:
            left, right = dsn.split("://", 1)
            if "@" in right and ":" in right.split("@", 1)[0]:
                creds, host = right.split("@", 1)
                user = creds.split(":", 1)[0]
                return f"{left}://{user}:***@{host}"
    except Exception:
        pass
    return dsn


def _infer_embedding_dim_from_db(conn: psycopg.Connection) -> int:
    with conn.cursor() as cur:
        cur.execute("SELECT embedding::text FROM rag_chunks LIMIT 1;")
        row = cur.fetchone()
    if not row or not row[0]:
        raise RuntimeError("Cannot infer embedding_dim: rag_chunks is empty.")
    s = str(row[0]).strip()
    if s.startswith("[") and s.endswith("]"):
        inner = s[1:-1].strip()
        if not inner:
            raise RuntimeError("Cannot infer embedding_dim: empty embedding literal.")
        return len([x for x in inner.split(",") if x.strip() != ""])
    raise RuntimeError(f"Cannot infer embedding_dim from embedding::text: {s[:80]}")


def _build_chunk_preview_rows(conn: psycopg.Connection, corpus_label: str = "") -> list[dict[str, Any]]:
    with conn.cursor() as cur:
        if corpus_label:
            cur.execute(
                """
                SELECT doc_path, chunk_index, chunk_text FROM rag_chunks
                WHERE corpus_label = %s
                ORDER BY doc_path, chunk_index;
                """,
                (corpus_label,),
            )
        else:
            cur.execute("SELECT doc_path, chunk_index, chunk_text FROM rag_chunks ORDER BY doc_path, chunk_index;")
        rows = cur.fetchall() or []
    preview: list[dict[str, Any]] = []
    for doc_path, chunk_index, chunk_text in rows:
        ct = chunk_text or ""
        preview.append({
            "doc_name": Path(str(doc_path)).name,
            "chunk_index": int(chunk_index),
            "chunk_chars": len(ct),
            "chunk_text": ct,
        })
    return preview


def _mean(nums: list) -> float | None:
    nums = [n for n in nums if n is not None]
    if not nums:
        return None
    return sum(nums) / float(len(nums))


def _pct(num: int, den: int) -> float | None:
    if den <= 0:
        return None
    return (float(num) / float(den)) * 100.0


def _cap_text(s: str, max_chars: int) -> str:
    s = s or ""
    return s if len(s) <= max_chars else (s[:max_chars] + "\n...[truncated]")


def _load_prompts(prompts_dir: Path) -> tuple[str, str, str, str]:
    """Returns (generator_system, generator_user_template, reviewer_system, reviewer_user_template)."""
    def _read(name: str) -> str:
        p = prompts_dir / name
        if not p.exists():
            raise RuntimeError(f"Missing prompt file: {p}")
        return p.read_text(encoding="utf-8")
    return (
        _read("generator_system.txt"),
        _read("generator_user.txt"),
        _read("reviewer_system.txt"),
        _read("reviewer_user.txt"),
    )


def _run_reviewer(
    gen_text: str,
    context_for_review: str,
    cfg_lm_url: str,
    cfg_review_model: str,
    reviewer_system: str,
    reviewer_user_template: str,
    temperature: float,
    max_tokens: int,
    timeout: int,
    review_provider: str = "lmstudio",
) -> tuple[dict[str, Any], float]:
    """Runs reviewer LLM and returns (rev_clean, elapsed_seconds)."""
    reviewer_user = (
        reviewer_user_template
        .replace("{{GEN_ITEM}}", gen_text)
        .replace("{{CONTEXT}}", context_for_review)
    )
    resolved_review_provider = "lmstudio" if review_provider == "local" else review_provider
    t0 = time.perf_counter()
    rev_raw = call_llm(
        lm_url=cfg_lm_url,
        model=cfg_review_model,
        system_prompt=reviewer_system,
        user_prompt=reviewer_user,
        temperature=temperature,
        max_tokens=max_tokens,
        request_timeout_seconds=timeout,
        provider=resolved_review_provider,
    )
    elapsed = time.perf_counter() - t0
    rev_json = extract_first_json_obj(rev_raw) or {}
    return enforce_hygiene_on_review(rev_json), elapsed


def _build_quality_meta(decisions_rows: list[dict[str, Any]]) -> dict[str, Any]:
    valid = [r for r in decisions_rows if bool(r.get("reviewer_schema_ok"))]
    sa_vals = [int(r["source_alignment"]) for r in valid if r.get("source_alignment") is not None]
    dq_vals = [int(r["distractor_quality"]) for r in valid if r.get("distractor_quality") is not None]
    sc_vals = [int(r["stem_clarity"]) for r in valid if r.get("stem_clarity") is not None]
    dm_vals = [bool(r["difficulty_match"]) for r in valid if r.get("difficulty_match") is not None]
    return {
        "quality.valid_review_rows": len(valid),
        "quality.mean_source_alignment": _mean(sa_vals),
        "quality.mean_distractor_quality": _mean(dq_vals),
        "quality.mean_stem_clarity": _mean(sc_vals),
        "quality.pct_source_alignment_gte_4": _pct(sum(1 for v in sa_vals if v >= 4), len(sa_vals)),
        "quality.pct_distractor_quality_gte_3": _pct(sum(1 for v in dq_vals if v >= 3), len(dq_vals)),
        "quality.pct_stem_clarity_gte_4": _pct(sum(1 for v in sc_vals if v >= 4), len(sc_vals)),
        "quality.pct_difficulty_match_true": _pct(sum(1 for v in dm_vals if v), len(dm_vals)),
    }


# ---------------------------------------------------------------------------
# RAG generation
# ---------------------------------------------------------------------------

def generate_from_db(cfg: GenerateConfig) -> dict[str, Any]:
    """RAG generate: retrieves top-k knowledge chunks, generates + reviews MCQ items."""
    prompts_dir = Path(cfg.prompts_dir)
    generator_system, generator_user_template, reviewer_system, reviewer_user_template = \
        _load_prompts(prompts_dir)

    condition_label = (os.environ.get("CONDITION_LABEL") or "rag").strip()
    created_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    out_dir = Path(cfg.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    manifest = {
        "created_at": created_at,
        "config": {
            "mode": "rag",
            "db_dsn_redacted": _redact_dsn(cfg.db_dsn),
            "lm_url": cfg.lm_url,
            "embed_model": cfg.embed_model,
            "generator_model": cfg.generator_model,
            "review_model": cfg.review_model,
            "n_items": int(cfg.n_items),
            "run_id": cfg.run_id,
            "prompts_dir": str(prompts_dir),
            "out_dir": str(out_dir),
            "top_k": int(cfg.top_k),
            "sleep_seconds": float(cfg.sleep_seconds),
            "condition_label": condition_label,
        },
    }
    manifest_path = out_dir / f"run_manifest_{cfg.run_id}.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    items_rows: list[dict[str, Any]] = []
    decisions_rows: list[dict[str, Any]] = []
    trace_rows: list[dict[str, Any]] = []
    pre_review_items: list[dict[str, Any]] = []  # for checkpoint 2

    schema_ok_count = 0
    reviewer_json_ok = 0
    reviewer_schema_ok_count = 0
    decisions_count: dict[str, int] = {}
    elapsed_gen = 0.0
    elapsed_review = 0.0

    corpus_label = (cfg.corpus_label or "").strip()

    with psycopg.connect(cfg.db_dsn) as conn:
        if chunks_rowcount(conn, corpus_label) <= 0:
            if corpus_label:
                raise RuntimeError(
                    "DB has 0 chunks for corpus_label=%r. Run ingest first." % corpus_label
                )
            raise RuntimeError("DB has 0 chunks. Run ingest first.")

        embedding_dim = _infer_embedding_dim_from_db(conn)
        ensure_schema(conn, embedding_dim)

        for i in range(int(cfg.n_items)):
            item_id = f"item_{i+1}"
            import datetime as _pdt
            print(
                "\n[%s] Generating %s..." % (_pdt.datetime.now().strftime("%H:%M:%S"), item_id),
                file=sys.stderr, flush=True,
            )

            seed_rows = get_random_chunks(conn, n=1, corpus_label=corpus_label)
            seed_text = seed_rows[0]["chunk_text"]
            seed_doc = seed_rows[0].get("doc_path", "")

            seed_emb = embed_texts(
                EmbedConfig(lm_url=cfg.lm_url, model=cfg.embed_model),
                [seed_text],
            )[0]

            retrieved = similarity_search(conn, seed_emb, int(cfg.top_k), corpus_label=corpus_label)

            for r in retrieved:
                trace_rows.append({
                    "run_id": cfg.run_id,
                    "item_id": item_id,
                    "doc_path": r.get("doc_path"),
                    "chunk_index": r.get("chunk_index"),
                    "distance": r.get("distance"),
                    "chunk_text": r.get("chunk_text"),
                })

            max_context_chars = int(os.environ.get("MAX_CONTEXT_CHARS_GEN", "3000"))
            context_parts = []
            for j, r in enumerate(retrieved):
                context_parts.append(f"[{j+1}] {r.get('chunk_text', '')}")
            context_block = _cap_text("\n\n".join(context_parts), max_context_chars)

            difficulty_target = (os.environ.get("DIFFICULTY_TARGET") or "medium").strip().lower()
            generator_user = (
                generator_user_template
                .replace("{{DIFFICULTY}}", difficulty_target)
                .replace("{{CONTEXT}}", context_block)
            )

            _gen_provider = "lmstudio" if cfg.generate_provider == "local" else cfg.generate_provider
            t0 = time.perf_counter()
            gen_raw = call_llm(
                lm_url=cfg.lm_url,
                model=cfg.generator_model,
                system_prompt=generator_system,
                user_prompt=generator_user,
                temperature=cfg.temperature_gen,
                max_tokens=cfg.max_tokens_gen,
                request_timeout_seconds=cfg.request_timeout_seconds,
                provider=_gen_provider,
            )
            item_elapsed = time.perf_counter() - t0
            elapsed_gen += item_elapsed
            print(
                "    [%s] Done in %.1fs" % (_pdt.datetime.now().strftime("%H:%M:%S"), item_elapsed),
                file=sys.stderr, flush=True,
            )
            gen_text = hard_trim_after_difficulty(clean_generator_text(gen_raw))
            schema_ok, violations = validate_generator_schema(gen_text)
            if schema_ok:
                schema_ok_count += 1

            parsed_fields = _parse_item_fields(gen_text)

            pre_review_items.append({
                "run_id": cfg.run_id,
                "item_id": item_id,
                "mode": "rag",
                "gen_text_clean": gen_text,
                "context_block": context_block,
                "seed_doc_path": seed_doc,
                "schema_ok": schema_ok,
                "schema_violations": "|".join(violations),
                "parsed_fields": parsed_fields,
                "human_edited": False,
            })

            if cfg.sleep_seconds:
                time.sleep(float(cfg.sleep_seconds))

        db_snap_summary = get_db_snapshot_summary(conn, corpus_label=corpus_label)
        db_snap_per_doc = get_db_snapshot_per_doc(conn, corpus_label=corpus_label)
        chunk_preview = _build_chunk_preview_rows(conn, corpus_label=corpus_label)

    # Checkpoint 2: human reviews items
    if cfg.checkpoint_items:
        pre_review_items = _checkpoint_items(pre_review_items)

    # Run reviewer on approved items
    for item in pre_review_items:
        item_id = item["item_id"]
        gen_text = item["gen_text_clean"]
        context_block = item.get("context_block", "")
        max_rev_chars = int(os.environ.get("MAX_CONTEXT_CHARS_REV", "2000"))
        context_for_review = _cap_text(context_block, max_rev_chars)

        rev_clean, elapsed = _run_reviewer(
            gen_text=gen_text,
            context_for_review=context_for_review,
            cfg_lm_url=cfg.lm_url,
            cfg_review_model=cfg.review_model,
            reviewer_system=reviewer_system,
            reviewer_user_template=reviewer_user_template,
            temperature=cfg.temperature_review,
            max_tokens=cfg.max_tokens_review,
            timeout=cfg.request_timeout_seconds,
            review_provider=cfg.review_provider,
        )
        elapsed_review += elapsed

        if rev_clean.get("reviewer_parse_ok"):
            reviewer_json_ok += 1
        if rev_clean.get("reviewer_schema_ok"):
            reviewer_schema_ok_count += 1

        decision = rev_clean.get("decision", "")
        decisions_count[decision] = decisions_count.get(decision, 0) + 1

        parsed_fields = item.get("parsed_fields", {})
        items_rows.append({
            "run_id": cfg.run_id,
            "item_id": item_id,
            "mode": "rag",
            "question": parsed_fields.get("question", ""),
            "a": parsed_fields.get("a", ""),
            "b": parsed_fields.get("b", ""),
            "c": parsed_fields.get("c", ""),
            "d": parsed_fields.get("d", ""),
            "correct_key": parsed_fields.get("correct_key", ""),
            "difficulty": parsed_fields.get("difficulty", ""),
            "decision": decision,
            "source_alignment": rev_clean.get("source_alignment"),
            "distractor_quality": rev_clean.get("distractor_quality"),
            "stem_clarity": rev_clean.get("stem_clarity"),
            "difficulty_match": rev_clean.get("difficulty_match"),
            "schema_ok": item.get("schema_ok"),
            "schema_violations": item.get("schema_violations"),
            "reviewer_schema_ok": bool(rev_clean.get("reviewer_schema_ok", False)),
            "reviewer_schema_violations": "|".join(rev_clean.get("reviewer_schema_violations", []) or []),
            "gen_text_clean": gen_text,
            "human_edited": item.get("human_edited", False),
            "seed_doc_path": item.get("seed_doc_path", ""),
        })

        decisions_rows.append({
            "run_id": cfg.run_id,
            "item_id": item_id,
            "decision": decision,
            "source_alignment": rev_clean.get("source_alignment"),
            "distractor_quality": rev_clean.get("distractor_quality"),
            "stem_clarity": rev_clean.get("stem_clarity"),
            "difficulty_match": rev_clean.get("difficulty_match"),
            "failure_layer": rev_clean.get("failure_layer", ""),
            "reason_codes": rev_clean.get("reason_codes", []),
            "revision_instructions": rev_clean.get("revision_instructions", ""),
            "reviewer_schema_ok": bool(rev_clean.get("reviewer_schema_ok", False)),
            "reviewer_schema_violations": rev_clean.get("reviewer_schema_violations", []),
            "reviewer_parse_ok": bool(rev_clean.get("reviewer_parse_ok", False)),
            "human_override": False,
            "human_corrected_text": "",
        })

    # Checkpoint 3: human reviews flagged items
    if cfg.checkpoint_review:
        decisions_rows = _checkpoint_review(decisions_rows)

    print(
        f"Timing: gen={elapsed_gen:.1f}s  review={elapsed_review:.1f}s",
        file=sys.stderr, flush=True,
    )

    quality_meta = _build_quality_meta(decisions_rows)
    meta = {
        "created_at": created_at,
        "run_id": cfg.run_id,
        "mode": "rag",
        "corpus_label": corpus_label,
        "condition_label": condition_label,
        "lm_url": cfg.lm_url,
        "embed_model": cfg.embed_model,
        "generator_model": cfg.generator_model,
        "review_model": cfg.review_model,
        "n_items": int(cfg.n_items),
        "top_k": int(cfg.top_k),
        "db_dsn_redacted": _redact_dsn(cfg.db_dsn),
        "timing.gen_seconds": round(elapsed_gen, 1),
        "timing.review_seconds": round(elapsed_review, 1),
        **quality_meta,
    }

    xlsx_path = write_run_xlsx(
        out_dir=out_dir,
        run_id=cfg.run_id,
        metadata=meta,
        db_snapshot_summary=db_snap_summary,
        db_snapshot_per_doc=db_snap_per_doc,
        chunk_preview_rows=chunk_preview,
        items_rows=items_rows,
        decisions_rows=decisions_rows,
        trace_rows=trace_rows,
        include_generation=True,
        include_chunk_preview=True,
    )

    return {
        "run_id": cfg.run_id,
        "mode": "rag",
        "condition_label": condition_label,
        "items_total": len(items_rows),
        "items_schema_ok": schema_ok_count,
        "reviewer_json_ok": reviewer_json_ok,
        "reviewer_schema_ok": reviewer_schema_ok_count,
        "decisions": decisions_count,
        "files": {"xlsx": str(xlsx_path), "manifest_json": str(manifest_path)},
    }


# ---------------------------------------------------------------------------
# Baseline (no-RAG) generation
# ---------------------------------------------------------------------------

def generate_baseline(cfg: BaselineConfig) -> dict[str, Any]:
    """Baseline mode: load docs directly, generate MCQ items without pgvector."""
    prompts_dir = Path(cfg.prompts_dir)
    generator_system, generator_user_template, reviewer_system, reviewer_user_template = \
        _load_prompts(prompts_dir)

    condition_label = (os.environ.get("CONDITION_LABEL") or "baseline").strip()
    created_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    out_dir = Path(cfg.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # Load all docs
    domain_dir = Path(cfg.domain_dir).resolve()
    docs: list[LoadedDoc] = []
    for p in sorted(domain_dir.rglob("*")):
        if p.is_file():
            doc = load_document(p)
            if doc is not None:
                docs.append(doc)

    if not docs:
        raise RuntimeError(f"No supported documents found in {domain_dir}")

    # Build full context from all docs (or a rotating subset per item)
    max_context_chars = int(os.environ.get("MAX_CONTEXT_CHARS_GEN", "6000"))
    all_doc_text = "\n\n".join(
        f"=== {doc.path.name} ===\n{doc.text}" for doc in docs
    )
    full_context = _cap_text(all_doc_text, max_context_chars)

    manifest = {
        "created_at": created_at,
        "config": {
            "mode": "baseline",
            "domain_dir": str(domain_dir),
            "lm_url": cfg.lm_url,
            "generator_model": cfg.generator_model,
            "review_model": cfg.review_model,
            "n_items": int(cfg.n_items),
            "run_id": cfg.run_id,
            "condition_label": condition_label,
        },
    }
    manifest_path = out_dir / f"run_manifest_{cfg.run_id}.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    items_rows: list[dict[str, Any]] = []
    decisions_rows: list[dict[str, Any]] = []
    pre_review_items: list[dict[str, Any]] = []

    schema_ok_count = 0
    reviewer_json_ok = 0
    reviewer_schema_ok_count = 0
    decisions_count: dict[str, int] = {}
    elapsed_gen = 0.0
    elapsed_review = 0.0

    for i in range(int(cfg.n_items)):
        item_id = f"item_{i+1}"
        print(f"Generating baseline {item_id}...", file=sys.stderr, flush=True)

        difficulty_target = (os.environ.get("DIFFICULTY_TARGET") or "medium").strip().lower()
        generator_user = (
            generator_user_template
            .replace("{{DIFFICULTY}}", difficulty_target)
            .replace("{{CONTEXT}}", full_context)
        )

        t0 = time.perf_counter()
        gen_raw = call_llm(
            lm_url=cfg.lm_url,
            model=cfg.generator_model,
            system_prompt=generator_system,
            user_prompt=generator_user,
            temperature=cfg.temperature_gen,
            max_tokens=cfg.max_tokens_gen,
            request_timeout_seconds=cfg.request_timeout_seconds,
            provider="lmstudio" if cfg.generate_provider == "local" else cfg.generate_provider,
        )
        elapsed_gen += time.perf_counter() - t0

        gen_text = hard_trim_after_difficulty(clean_generator_text(gen_raw))
        schema_ok, violations = validate_generator_schema(gen_text)
        if schema_ok:
            schema_ok_count += 1
        parsed_fields = _parse_item_fields(gen_text)

        pre_review_items.append({
            "run_id": cfg.run_id,
            "item_id": item_id,
            "mode": "baseline",
            "gen_text_clean": gen_text,
            "context_block": full_context,
            "seed_doc_path": "",
            "schema_ok": schema_ok,
            "schema_violations": "|".join(violations),
            "parsed_fields": parsed_fields,
            "human_edited": False,
        })

        if cfg.sleep_seconds:
            time.sleep(float(cfg.sleep_seconds))

    # Checkpoint 2
    if cfg.checkpoint_items:
        pre_review_items = _checkpoint_items(pre_review_items)

    for item in pre_review_items:
        item_id = item["item_id"]
        gen_text = item["gen_text_clean"]
        context_for_review = _cap_text(item.get("context_block", ""), 2000)

        rev_clean, elapsed = _run_reviewer(
            gen_text=gen_text,
            context_for_review=context_for_review,
            cfg_lm_url=cfg.lm_url,
            cfg_review_model=cfg.review_model,
            reviewer_system=reviewer_system,
            reviewer_user_template=reviewer_user_template,
            temperature=cfg.temperature_review,
            max_tokens=cfg.max_tokens_review,
            timeout=cfg.request_timeout_seconds,
            review_provider=cfg.review_provider,
        )
        elapsed_review += elapsed

        if rev_clean.get("reviewer_parse_ok"):
            reviewer_json_ok += 1
        if rev_clean.get("reviewer_schema_ok"):
            reviewer_schema_ok_count += 1

        decision = rev_clean.get("decision", "")
        decisions_count[decision] = decisions_count.get(decision, 0) + 1

        parsed_fields = item.get("parsed_fields", {})
        items_rows.append({
            "run_id": cfg.run_id,
            "item_id": item_id,
            "mode": "baseline",
            "question": parsed_fields.get("question", ""),
            "a": parsed_fields.get("a", ""),
            "b": parsed_fields.get("b", ""),
            "c": parsed_fields.get("c", ""),
            "d": parsed_fields.get("d", ""),
            "correct_key": parsed_fields.get("correct_key", ""),
            "difficulty": parsed_fields.get("difficulty", ""),
            "decision": decision,
            "source_alignment": rev_clean.get("source_alignment"),
            "distractor_quality": rev_clean.get("distractor_quality"),
            "stem_clarity": rev_clean.get("stem_clarity"),
            "difficulty_match": rev_clean.get("difficulty_match"),
            "schema_ok": item.get("schema_ok"),
            "schema_violations": item.get("schema_violations"),
            "reviewer_schema_ok": bool(rev_clean.get("reviewer_schema_ok", False)),
            "reviewer_schema_violations": "|".join(rev_clean.get("reviewer_schema_violations", []) or []),
            "gen_text_clean": gen_text,
            "human_edited": item.get("human_edited", False),
            "seed_doc_path": "",
        })

        decisions_rows.append({
            "run_id": cfg.run_id,
            "item_id": item_id,
            "decision": decision,
            "source_alignment": rev_clean.get("source_alignment"),
            "distractor_quality": rev_clean.get("distractor_quality"),
            "stem_clarity": rev_clean.get("stem_clarity"),
            "difficulty_match": rev_clean.get("difficulty_match"),
            "failure_layer": rev_clean.get("failure_layer", ""),
            "reason_codes": rev_clean.get("reason_codes", []),
            "revision_instructions": rev_clean.get("revision_instructions", ""),
            "reviewer_schema_ok": bool(rev_clean.get("reviewer_schema_ok", False)),
            "reviewer_schema_violations": rev_clean.get("reviewer_schema_violations", []),
            "reviewer_parse_ok": bool(rev_clean.get("reviewer_parse_ok", False)),
            "human_override": False,
            "human_corrected_text": "",
        })

    # Checkpoint 3
    if cfg.checkpoint_review:
        decisions_rows = _checkpoint_review(decisions_rows)

    quality_meta = _build_quality_meta(decisions_rows)
    meta = {
        "created_at": created_at,
        "run_id": cfg.run_id,
        "mode": "baseline",
        "condition_label": condition_label,
        "domain_dir": str(domain_dir),
        "lm_url": cfg.lm_url,
        "generator_model": cfg.generator_model,
        "review_model": cfg.review_model,
        "n_items": int(cfg.n_items),
        "timing.gen_seconds": round(elapsed_gen, 1),
        "timing.review_seconds": round(elapsed_review, 1),
        **quality_meta,
    }

    xlsx_path = write_run_xlsx(
        out_dir=out_dir,
        run_id=cfg.run_id,
        metadata=meta,
        chunk_preview_rows=None,
        items_rows=items_rows,
        decisions_rows=decisions_rows,
        trace_rows=None,
        include_generation=True,
        include_chunk_preview=False,
    )

    return {
        "run_id": cfg.run_id,
        "mode": "baseline",
        "condition_label": condition_label,
        "items_total": len(items_rows),
        "items_schema_ok": schema_ok_count,
        "reviewer_json_ok": reviewer_json_ok,
        "reviewer_schema_ok": reviewer_schema_ok_count,
        "decisions": decisions_count,
        "files": {"xlsx": str(xlsx_path), "manifest_json": str(manifest_path)},
    }


# ---------------------------------------------------------------------------
# Pipeline orchestrator
# ---------------------------------------------------------------------------

def run_pipeline(cfg: PipelineConfig) -> dict[str, Any]:
    """Orchestrates ingest (with optional chunk checkpoint) and generate."""
    ingest_ran = False
    ingest_summary: dict[str, Any] | None = None
    chunks_by_doc: dict[str, list[str]] = {}

    prompts_dir = Path(cfg.prompts_dir)

    # Derive corpus_label from domain_dir basename when not explicitly set.
    corpus_label = (cfg.corpus_label or "").strip() or Path(cfg.domain_dir).resolve().name

    with psycopg.connect(cfg.db_dsn) as conn:
        migrate_corpus_label(conn)
        has_chunks = chunks_rowcount(conn, corpus_label) > 0

    if cfg.force_ingest or not has_chunks:
        ingest_ran = True
        ingest_cfg = IngestConfig(
            domain_dir=Path(cfg.domain_dir),
            db_dsn=cfg.db_dsn,
            embed_lm_url=cfg.lm_url,
            embed_model=cfg.embed_model,
            lm_url=cfg.lm_url,
            context_model=cfg.context_model,
            api_provider=cfg.api_provider,
            api_model=cfg.api_model,
            ingest_provider=cfg.ingest_provider,
            corpus_label=corpus_label,
            embedding_dim=int(cfg.embedding_dim) if cfg.embedding_dim is not None else None,
            batch_size=int(cfg.batch_size),
            clear_first=bool(cfg.clear_first),
            ingest_delay_seconds=float(cfg.ingest_delay_seconds),
        )
        ingest_t0 = time.perf_counter()
        ingest_summary = ingest_domain(ingest_cfg, prompts_dir)
        ingest_elapsed = time.perf_counter() - ingest_t0
        print(
            "\nIngest total time: %dm %02ds" % (int(ingest_elapsed) // 60, int(ingest_elapsed) % 60),
            file=sys.stderr, flush=True,
        )
        # Checkpoint 1: review chunks just written to DB
        if cfg.checkpoint_chunks:
            with psycopg.connect(cfg.db_dsn) as _cp1_conn:
                _cp1_rows = _build_chunk_preview_rows(_cp1_conn, corpus_label=corpus_label)
            _cp1_by_doc = {}
            for r in _cp1_rows:
                key = r.get("rel_path") or r.get("doc_path", "unknown")
                _cp1_by_doc.setdefault(key, []).append(r.get("chunk_text", ""))
            _checkpoint_chunks(_cp1_by_doc)

    # Ingest-only path
    if cfg.ingest_only:
        created_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        out_dir = Path(cfg.out_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        run_id = cfg.run_id or datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%SZ")

        meta = {
            "created_at": created_at,
            "run_id": run_id,
            "mode": "ingest_only",
            "corpus_label": corpus_label,
            "lm_url": cfg.lm_url,
            "embed_model": cfg.embed_model,
            "context_model": cfg.context_model,
            "db_dsn_redacted": _redact_dsn(cfg.db_dsn),
            "domain_dir": str(cfg.domain_dir),
            "ingest_summary": json.dumps(ingest_summary or {}, ensure_ascii=False),
        }

        with psycopg.connect(cfg.db_dsn) as conn:
            db_snap_summary = get_db_snapshot_summary(conn, corpus_label=corpus_label)
            db_snap_per_doc = get_db_snapshot_per_doc(conn, corpus_label=corpus_label)
            chunk_preview = _build_chunk_preview_rows(conn, corpus_label=corpus_label)

        manifest = {"created_at": created_at, "config": meta}
        manifest_path = out_dir / f"run_manifest_{run_id}.json"
        manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

        xlsx_path = write_run_xlsx(
            out_dir=out_dir,
            run_id=run_id,
            metadata=meta,
            db_snapshot_summary=db_snap_summary,
            db_snapshot_per_doc=db_snap_per_doc,
            chunk_preview_rows=chunk_preview,
            include_generation=False,
            include_chunk_preview=True,
        )

        print(f"Ingest-only XLSX written: {xlsx_path}", file=sys.stderr, flush=True)
        return {
            "ingest_ran": ingest_ran,
            "ingest_summary": ingest_summary,
            "ingest_only": True,
            "corpus_label": corpus_label,
            "files": {"xlsx": str(xlsx_path), "manifest_json": str(manifest_path)},
        }

    # Full pipeline: generate (RAG mode)
    gen_cfg = GenerateConfig(
        db_dsn=cfg.db_dsn,
        lm_url=cfg.lm_url,
        embed_model=cfg.embed_model,
        generator_model=cfg.generator_model,
        review_model=cfg.review_model,
        n_items=int(cfg.n_items),
        run_id=cfg.run_id,
        prompts_dir=prompts_dir,
        out_dir=Path(cfg.out_dir),
        generate_provider=cfg.generate_provider,
        review_provider=cfg.review_provider,
        corpus_label=corpus_label,
        top_k=int(cfg.top_k),
        sleep_seconds=float(cfg.sleep_seconds),
        checkpoint_items=bool(cfg.checkpoint_items),
        checkpoint_review=bool(cfg.checkpoint_review),
    )

    generate_summary = generate_from_db(gen_cfg)
    return {
        "ingest_ran": ingest_ran,
        "ingest_summary": ingest_summary,
        "corpus_label": corpus_label,
        "generate_summary": generate_summary,
    }
