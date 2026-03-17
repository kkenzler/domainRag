"""assess_run.py — Programmatic quality check for a completed run XLSX.

Usage: python assess_run.py <path_to_xlsx>

Prints a structured summary and exits 0 (pass) or 1 (fail).
Pass criteria:
  - >= 80% items schema-valid
  - >= 60% reviewer ACCEPT
  - Difficulty labels in output match DIFFICULTY_TARGET env var (>= 60%)
"""

from __future__ import annotations

import io
import os
import sys
from pathlib import Path

# Force UTF-8 output on Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed")
    sys.exit(2)


PASS_SCHEMA_RATE = 0.80
PASS_ACCEPT_RATE = 0.60
PASS_DIFFICULTY_RATE = 0.60


def _col_index(sheet, header_name):
    for cell in sheet[1]:
        if cell.value and str(cell.value).strip().lower() == header_name.lower():
            return cell.column
    return None


def assess(xlsx_path: str) -> bool:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    difficulty_target = (os.environ.get("DIFFICULTY_TARGET") or "").strip().lower()

    # ---- Items sheet --------------------------------------------------------
    if "Items" not in wb.sheetnames:
        print("FAIL: no Items sheet found")
        return False

    items_ws = wb["Items"]
    rows = list(items_ws.iter_rows(values_only=True))
    if len(rows) < 2:
        print("FAIL: Items sheet has no data rows")
        return False

    headers = [str(c).strip() if c else "" for c in rows[0]]
    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    schema_col = col("schema_ok")
    diff_col = col("difficulty")
    item_id_col = col("item_id")
    gen_text_col = col("gen_text_clean")

    data_rows = rows[1:]
    total = len(data_rows)

    schema_ok = sum(1 for r in data_rows if schema_col is not None and str(r[schema_col]).strip().lower() in ("true", "1", "yes")) if schema_col is not None else total
    schema_rate = schema_ok / total if total else 0

    difficulty_matches = 0
    print("\n--- Items ---")
    for r in data_rows:
        item_id = r[item_id_col] if item_id_col is not None else "?"
        diff = str(r[diff_col]).strip().lower() if diff_col is not None and r[diff_col] else "?"
        gen_text = str(r[gen_text_col])[:300] if gen_text_col is not None and r[gen_text_col] else ""
        match = (diff == difficulty_target) if difficulty_target and difficulty_target != "any" else True
        if match:
            difficulty_matches += 1
        flag = "" if match else "  [DIFFICULTY MISMATCH]"
        print(f"  {item_id} | difficulty={diff}{flag}")
        print(f"    {gen_text[:200]}")
        print()

    diff_rate = difficulty_matches / total if total else 0

    # ---- Reviewer Decisions sheet ------------------------------------------
    accept_count = 0
    rev_total = 0
    if "Reviewer Decisions" in wb.sheetnames:
        rev_ws = wb["Reviewer Decisions"]
        rev_rows = list(rev_ws.iter_rows(values_only=True))
        if len(rev_rows) >= 2:
            rev_headers = [str(c).strip() if c else "" for c in rev_rows[0]]
            def rcol(name):
                try:
                    return rev_headers.index(name)
                except ValueError:
                    return None
            dec_col = rcol("decision")
            rc_col = rcol("reason_codes")
            ri_col = rcol("revision_instructions")
            print("--- Reviewer Decisions ---")
            for r in rev_rows[1:]:
                rev_total += 1
                decision = str(r[dec_col]).strip() if dec_col is not None and r[dec_col] else "?"
                if decision == "ACCEPT":
                    accept_count += 1
                reason = str(r[rc_col]) if rc_col is not None and r[rc_col] else ""
                instr = str(r[ri_col])[:150] if ri_col is not None and r[ri_col] else ""
                print(f"  {decision}  reasons={reason}")
                if instr:
                    print(f"    -> {instr}")
            print()

    accept_rate = accept_count / rev_total if rev_total else 0

    # ---- Summary -----------------------------------------------------------
    print("--- Summary ---")
    print("  total_items      : %d" % total)
    print("  schema_ok        : %d / %d  (%.0f%%)" % (schema_ok, total, schema_rate * 100))
    print("  reviewer_accept  : %d / %d  (%.0f%%)" % (accept_count, rev_total, accept_rate * 100))
    if difficulty_target and difficulty_target != "any":
        print("  difficulty_match : %d / %d  (%.0f%%)  [target=%s]" % (difficulty_matches, total, diff_rate * 100, difficulty_target))

    passed = (
        schema_rate >= PASS_SCHEMA_RATE and
        accept_rate >= PASS_ACCEPT_RATE and
        (diff_rate >= PASS_DIFFICULTY_RATE if difficulty_target and difficulty_target != "any" else True)
    )

    print("\n  VERDICT: %s" % ("PASS" if passed else "FAIL"))
    return passed


if __name__ == "__main__":
    if len(sys.argv) < 2:
        # Auto-find most recent xlsx in runs/
        runs_dir = Path(__file__).parent / "runs"
        xls = sorted(runs_dir.glob("run_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if not xls:
            print("No run xlsx found in runs/")
            sys.exit(2)
        target = str(xls[0])
        print("Auto-selected: %s" % target)
    else:
        target = sys.argv[1]

    ok = assess(target)
    sys.exit(0 if ok else 1)
